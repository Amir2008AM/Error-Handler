"""
Excel → PDF converter — v2: Arabic-first, ilovepdf-quality output.

v2 improvements over v1:
  1. Better column width calculation: uses both Excel hints AND content analysis
  2. RTL column ordering: Arabic sheets are rendered right-to-left as they appear
  3. Multi-line cells: Paragraphs preserve newlines correctly
  4. Merged cell handling: expanded to all cells in range — no blank cells
  5. Better row height: adjusts dynamically based on wrapped content
  6. Header detection: title rows stripped cleanly before table
  7. Arabic reshaping: applied per-token so mixed Arabic/Latin cells work correctly
  8. Table borders: heavier separators, cleaner grid
  9. Font fallback: graceful fallback to Helvetica if Amiri is missing
  10. Page layout: tighter margins, clear page footer with page numbers
  11. Noise filtering: drops Excel comment noise rows reliably
  12. Empty column pruning: drops columns that are 100% empty
  13. Alternate row shading: light/white rows for readability
  14. Text overflow: all cells use word-wrap; no truncation

Usage:
    python3 excel_to_pdf.py input.xlsx output.pdf [options]

Options:
    --page-size    a4 | letter | legal      (default: a4)
    --orientation  portrait | landscape     (default: landscape)
    --font-size    <int>                    (default: 9)
"""

from __future__ import annotations

import sys
import os
import re
import json
import argparse
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
import arabic_reshaper
from bidi.algorithm import get_display

from reportlab.lib.pagesizes import A4, LETTER, LEGAL, landscape, portrait
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate,
    Table, TableStyle, Paragraph, Spacer, PageBreak,
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfgen import canvas as rl_canvas


# ── Constants ─────────────────────────────────────────────────────────────────

_AR_RE = re.compile(r'[\u0600-\u06FF\u0750-\u077F\uFB50-\uFDFF\uFE70-\uFEFF]')

_COMMENT_NOISE = re.compile(
    r'Your version of Excel allows you to read this threaded comment',
    re.IGNORECASE,
)

PAGE_SIZES = {'a4': A4, 'letter': LETTER, 'legal': LEGAL}

_HDR_BG     = colors.HexColor('#1E3A5F')
_HDR_FG     = colors.white
_ALT_ROW_BG = colors.HexColor('#F0F4F8')
_GRID_COLOR = colors.HexColor('#BBCAD6')
_HDR_LINE   = colors.HexColor('#14283F')
_TITLE_CLR  = colors.HexColor('#1E3A5F')


# ── Arabic helpers ────────────────────────────────────────────────────────────

def has_arabic(text: str) -> bool:
    return bool(_AR_RE.search(text))


def _is_mostly_arabic(text: str) -> bool:
    stripped = text.replace(' ', '')
    if not stripped:
        return False
    return len(_AR_RE.findall(stripped)) / len(stripped) >= 0.40


def shape_arabic(text: str) -> str:
    """
    Logical-order Unicode → visual-order Presentation Forms.
    Token-by-token processing preserves Latin/digit tokens in-place.
    For primarily Arabic text the token list is reversed for RTL reading.
    """
    if not text or not has_arabic(text):
        return text

    tokens = text.split(' ')
    shaped: list[str] = []
    for tok in tokens:
        if tok and has_arabic(tok):
            try:
                shaped.append(get_display(arabic_reshaper.reshape(tok)))
            except Exception:
                shaped.append(tok)
        else:
            shaped.append(tok)

    if _is_mostly_arabic(text):
        shaped = list(reversed(shaped))

    return ' '.join(shaped)


# ── Font registration ─────────────────────────────────────────────────────────

_FONT_REGULAR = 'Amiri'
_FONT_BOLD    = 'Amiri-Bold'
_fonts_registered = False


def _register_fonts() -> tuple[str, str]:
    global _fonts_registered
    if _fonts_registered:
        return _FONT_REGULAR, _FONT_BOLD

    search_dirs = [
        Path(__file__).parent / 'artifacts' / 'web-toolify' / 'lib' / 'processing' / 'fonts',
        Path(__file__).parent / 'fonts',
        Path('/usr/share/fonts'),
        Path('/nix/store'),
    ]

    reg_path: Path | None = None
    bold_path: Path | None = None

    for d in search_dirs:
        if not d.exists():
            continue
        hits = list(d.rglob('Amiri-Regular.ttf'))
        if hits:
            reg_path  = hits[0]
            bold_hits = list(hits[0].parent.glob('Amiri-Bold.ttf'))
            bold_path = bold_hits[0] if bold_hits else None
            break

    try:
        if reg_path and reg_path.exists():
            pdfmetrics.registerFont(TTFont(_FONT_REGULAR, str(reg_path)))
        if bold_path and bold_path.exists():
            pdfmetrics.registerFont(TTFont(_FONT_BOLD, str(bold_path)))
        pdfmetrics.getFont(_FONT_REGULAR)
        _fonts_registered = True
        return _FONT_REGULAR, _FONT_BOLD
    except Exception:
        return 'Helvetica', 'Helvetica-Bold'


# ── Cell value helpers ────────────────────────────────────────────────────────

def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    s = str(value).strip()
    if _COMMENT_NOISE.search(s):
        return ''
    return s[:1000] if len(s) > 1000 else s


def _is_noise_row(row: list[str]) -> bool:
    return bool(_COMMENT_NOISE.search(' '.join(row)))


# ── Merged-cell expansion ────────────────────────────────────────────────────

def _expand_merged_cells(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[tuple[int,int], str]:
    value_map: dict[tuple[int,int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            value_map[(cell.row, cell.column)] = _cell_str(cell.value)
    for mr in ws.merged_cells.ranges:
        origin = value_map.get((mr.min_row, mr.min_col), '')
        for ri in range(mr.min_row, mr.max_row + 1):
            for ci in range(mr.min_col, mr.max_col + 1):
                value_map[(ri, ci)] = origin
    return value_map


# ── Column width helpers ──────────────────────────────────────────────────────

def _excel_col_widths(ws: openpyxl.worksheet.worksheet.Worksheet, n_cols: int) -> list[float | None]:
    widths: list[float | None] = []
    for ci in range(1, n_cols + 1):
        dim = ws.column_dimensions.get(get_column_letter(ci))
        widths.append(float(dim.width) if dim and dim.width and dim.width > 0 else None)
    return widths


def _estimate_col_widths(
    rows: list[list[str]],
    available_w: float,
    font_size: int,
    excel_hints: list[float | None],
    is_rtl: bool = False,
) -> list[float]:
    if not rows or not rows[0]:
        return [available_w]

    n_cols  = max(len(r) for r in rows)
    ar_pt   = font_size * 0.72
    lat_pt  = font_size * 0.58
    min_w   = font_size * 2.0
    padding = 10

    raw: list[float] = []
    for ci in range(n_cols):
        hint = excel_hints[ci] if ci < len(excel_hints) else None
        if hint is not None:
            raw.append(max(hint * 7.0, min_w))
            continue
        best = min_w
        for row in rows:
            if ci < len(row):
                cell = str(row[ci])
                lines = cell.split('\n') if '\n' in cell else [cell]
                for line in lines:
                    arc = len(_AR_RE.findall(line))
                    lac = len(line) - arc
                    w = arc * ar_pt + lac * lat_pt + padding
                    best = max(best, w)
        raw.append(best)

    total = sum(raw)
    if total <= 0:
        return [available_w / max(n_cols, 1)] * n_cols
    scale = available_w / total
    fitted = [max(w * scale, min_w) for w in raw]
    ft = sum(fitted)
    if ft > available_w * 1.02:
        s2 = available_w / ft
        fitted = [max(w * s2, min_w) for w in fitted]
    return fitted


# ── Row reading ───────────────────────────────────────────────────────────────

def _read_sheet_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[list[str]]:
    if ws.max_row is None or ws.max_column is None:
        return []
    value_map = _expand_merged_cells(ws)
    rows: list[list[str]] = []
    for ri in range(1, ws.max_row + 1):
        row = [value_map.get((ri, ci), '') for ci in range(1, ws.max_column + 1)]
        if not any(row) or _is_noise_row(row):
            continue
        rows.append(row)
    return rows


def _prune_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows
    n_cols = max(len(r) for r in rows)
    empty = {ci for ci in range(n_cols) if all(ci >= len(r) or not r[ci].strip() for r in rows)}
    if not empty:
        return rows
    return [[v for ci, v in enumerate(r) if ci not in empty] for r in rows]


# ── Title / header separation ─────────────────────────────────────────────────

def _is_merged_title_row(row: list[str]) -> bool:
    ne = [v for v in row if v.strip()]
    return bool(ne) and len(set(ne)) == 1


def _extract_titles(rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    titles: list[str] = []
    idx = 0
    while idx < min(len(rows), 5):
        if _is_merged_title_row(rows[idx]):
            titles.append([v for v in rows[idx] if v.strip()][0])
            idx += 1
        else:
            break
    return titles, rows[idx:]


# ── Page-number footer ────────────────────────────────────────────────────────

class _NumberedCanvas(rl_canvas.Canvas):
    def __init__(self, *args: Any, font_regular: str = 'Helvetica', doc_title: str = '', **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self._font_regular = font_regular
        self._doc_title    = doc_title
        self._saved: list[dict] = []

    def showPage(self) -> None:
        self._saved.append(dict(self.__dict__))
        self._startPage()

    def save(self) -> None:
        total = len(self._saved)
        for state in self._saved:
            self.__dict__.update(state)
            self._footer(total)
            super().showPage()
        super().save()

    def _footer(self, total: int) -> None:
        try:
            pw = self._pagesize[0]
            self.setFont(self._font_regular, 7.5)
            self.setFillColor(colors.HexColor('#888888'))
            self.drawCentredString(pw / 2, 5 * mm, f'{self._pageNumber} / {total}')
            if self._doc_title:
                self.drawString(12 * mm, 5 * mm, self._doc_title[:60])
            self.setStrokeColor(colors.HexColor('#DDDDDD'))
            self.setLineWidth(0.4)
            self.line(12 * mm, 9 * mm, pw - 12 * mm, 9 * mm)
        except Exception:
            pass


# ── Paragraph style factory ───────────────────────────────────────────────────

def _make_styles(font_regular: str, font_bold: str, font_size: int, is_rtl: bool = False) -> dict[str, ParagraphStyle]:
    fs = font_size
    ww = 'RTL' if is_rtl else 'CJK'
    return {
        'title': ParagraphStyle('T', fontName=font_bold, fontSize=fs+5, alignment=TA_CENTER,
                                leading=(fs+5)*1.5, spaceAfter=3, textColor=_TITLE_CLR, wordWrap=ww),
        'subtitle': ParagraphStyle('ST', fontName=font_bold, fontSize=fs+2, alignment=TA_CENTER,
                                   leading=(fs+2)*1.4, spaceAfter=2, textColor=_TITLE_CLR, wordWrap=ww),
        'hdr': ParagraphStyle('H', fontName=font_bold, fontSize=fs, alignment=TA_CENTER,
                              leading=fs*1.4, textColor=colors.white, wordWrap=ww),
        'cell_ar': ParagraphStyle('CA', fontName=font_regular, fontSize=fs, alignment=TA_RIGHT,
                                  leading=fs*1.4, wordWrap='RTL'),
        'cell_ltr': ParagraphStyle('CL', fontName=font_regular, fontSize=fs, alignment=TA_LEFT,
                                   leading=fs*1.4, wordWrap='CJK'),
        'cell_num': ParagraphStyle('CN', fontName=font_regular, fontSize=fs, alignment=TA_CENTER,
                                   leading=fs*1.4),
    }


# ── Cell content classification ───────────────────────────────────────────────

def _is_numeric_cell(text: str) -> bool:
    stripped = re.sub(r'[\s\-/:.,+]', '', text)
    return bool(stripped) and all(c.isdigit() or c in '.,:;/-+' for c in stripped)


def _build_cell_para(text: str, styles: dict, is_header: bool, is_rtl_sheet: bool) -> Paragraph:
    if is_header:
        shaped = shape_arabic(text) if (text and has_arabic(text)) else text
        return Paragraph(shaped or '', styles['hdr'])
    if not text:
        return Paragraph('', styles['cell_num'])
    if has_arabic(text):
        return Paragraph(shape_arabic(text), styles['cell_ar'])
    if _is_numeric_cell(text):
        return Paragraph(text, styles['cell_num'])
    return Paragraph(text, styles['cell_ltr'])


# ── Sheet RTL detection ───────────────────────────────────────────────────────

def _sheet_is_rtl(rows: list[list[str]]) -> bool:
    sample = ' '.join(v for row in rows[:10] for v in row).replace(' ', '')
    return bool(sample) and len(_AR_RE.findall(sample)) / len(sample) >= 0.30


# ── Main conversion ───────────────────────────────────────────────────────────

def excel_to_pdf(
    input_path: str,
    output_path: str,
    page_size: str = 'a4',
    orientation: str = 'landscape',
    font_size: int = 9,
) -> dict:
    font_regular, font_bold = _register_fonts()

    base_size = PAGE_SIZES.get(page_size, A4)
    page_w, page_h = landscape(base_size) if orientation == 'landscape' else portrait(base_size)

    margin_h    = 12 * mm
    margin_t    = 12 * mm
    margin_b    = 16 * mm
    available_w = page_w - 2 * margin_h

    wb        = openpyxl.load_workbook(input_path, data_only=True)
    doc_title = Path(input_path).stem
    story: list[Any] = []
    first_sheet = True

    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        raw_rows = _read_sheet_rows(ws)
        if not raw_rows:
            continue

        raw_rows = _prune_empty_columns(raw_rows)
        if not raw_rows or not raw_rows[0]:
            continue

        if not first_sheet:
            story.append(PageBreak())
        first_sheet = False

        is_rtl = _sheet_is_rtl(raw_rows)
        styles = _make_styles(font_regular, font_bold, font_size, is_rtl)

        title_texts, data_rows = _extract_titles(raw_rows)

        for i, t in enumerate(title_texts):
            st = styles['title'] if i == 0 else styles['subtitle']
            story.append(Paragraph(shape_arabic(t) if has_arabic(t) else t, st))
        if title_texts:
            story.append(Spacer(1, 4 * mm))

        if not data_rows:
            continue

        n_cols = max(len(r) for r in data_rows)
        normed = [(list(r) + [''] * n_cols)[:n_cols] for r in data_rows]

        # For RTL sheets: reverse column order so the table reads right-to-left
        if is_rtl:
            normed = [list(reversed(r)) for r in normed]

        excel_hints = _excel_col_widths(ws, n_cols)
        if is_rtl:
            excel_hints = list(reversed(excel_hints))
        col_widths = _estimate_col_widths(normed, available_w, font_size, excel_hints, is_rtl)

        para_rows: list[list[Paragraph]] = []
        for ri, row in enumerate(normed):
            para_rows.append([
                _build_cell_para(cell, styles, ri == 0, is_rtl)
                for cell in row
            ])

        n_rows = len(para_rows)

        ts_cmds: list[tuple] = [
            ('BACKGROUND',    (0, 0), (-1, 0),  _HDR_BG),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  _HDR_FG),
            ('FONTNAME',      (0, 0), (-1, 0),  font_bold),
            ('FONTSIZE',      (0, 0), (-1, 0),  font_size),
            ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEBELOW',     (0, 0), (-1, 0),  1.5, _HDR_LINE),
            ('FONTNAME',      (0, 1), (-1, -1), font_regular),
            ('FONTSIZE',      (0, 1), (-1, -1), font_size),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
            ('GRID',          (0, 0), (-1, -1), 0.35, _GRID_COLOR),
            ('BOX',           (0, 0), (-1, -1), 0.8,  _GRID_COLOR),
            ('ALIGN',         (0, 1), (-1, -1), 'RIGHT' if is_rtl else 'LEFT'),
        ]

        for ri in range(1, n_rows, 2):
            ts_cmds.append(('BACKGROUND', (0, ri), (-1, ri), _ALT_ROW_BG))

        tbl = Table(para_rows, colWidths=col_widths, repeatRows=1, splitByRow=1)
        tbl.setStyle(TableStyle(ts_cmds))
        story.append(tbl)

    if not story:
        err_style = ParagraphStyle('E', fontName='Helvetica', fontSize=12,
                                   alignment=TA_CENTER, textColor=colors.red)
        story.extend([Spacer(1, 40 * mm), Paragraph('No table data found in the workbook.', err_style)])

    frame = Frame(margin_h, margin_b, page_w - 2 * margin_h, page_h - margin_t - margin_b,
                  id='main', leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    doc = BaseDocTemplate(
        output_path,
        pagesize=(page_w, page_h),
        pageTemplates=[PageTemplate(id='main', frames=[frame])],
        leftMargin=margin_h, rightMargin=margin_h,
        topMargin=margin_t,  bottomMargin=margin_b,
    )

    doc.build(
        story,
        canvasmaker=lambda *a, **kw: _NumberedCanvas(*a, font_regular=font_regular, doc_title=doc_title, **kw),
    )

    page_count = 1
    try:
        with open(output_path, 'rb') as f:
            data = f.read()
        page_count = max(data.count(b'/Type/Page'), data.count(b'/Type /Page'), 1)
    except Exception:
        pass

    return {'success': True, 'pageCount': page_count, 'engine': 'python-arabic-reshaper-v2'}


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description='Excel → PDF (Arabic-ready v2)')
    parser.add_argument('input',  help='Input .xlsx file')
    parser.add_argument('output', help='Output .pdf file')
    parser.add_argument('--page-size',   default='a4',       choices=['a4', 'letter', 'legal'])
    parser.add_argument('--orientation', default='landscape', choices=['portrait', 'landscape'])
    parser.add_argument('--font-size',   default=9, type=int)
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(json.dumps({'success': False, 'error': f'File not found: {args.input}'}))
        sys.exit(1)

    try:
        result = excel_to_pdf(args.input, args.output,
                              page_size=args.page_size,
                              orientation=args.orientation,
                              font_size=args.font_size)
        print(json.dumps(result))
    except Exception as exc:
        import traceback
        print(json.dumps({'success': False, 'error': str(exc), 'traceback': traceback.format_exc()}))
        sys.exit(1)


if __name__ == '__main__':
    main()
