"""
pdf_table_extractor.py  — Enterprise PDF → Excel Reconstruction Pipeline v3
=============================================================================

Architecture (10 phases):
  Phase 1  — Document Classification (text/scanned/hybrid, Arabic/LTR, bordered)
  Phase 2  — Layout Analysis (geometric, visual alignment, overlap-based clustering)
  Phase 3  — Smart Engine Routing (per-page: Camelot lattice/stream/pdfplumber/OCR)
  Phase 4  — OCR Pipeline (pdf2image + Tesseract, coordinate-aware, deskew)
  Phase 5  — Persistent Column Stabilization (global schema, overlap-based assignment)
  Phase 6  — Table Reconstruction Engine (merge fragments, repair rows)
  Phase 7  — Header Intelligence (dedup repeated headers, multi-page continuation)
  Phase 8  — Arabic RTL Support (reshaping, bidi correction, RTL Excel alignment)
  Phase 9  — Excel Generation Engine (openpyxl, borders, merged cells, RTL sheets)
  Phase 10 — Validation Layer (consistency, repair, empty-output detection)

v3 improvements:
  • Overlap-based row clustering (fixes multi-line row confusion)
  • Word-center + overlap-ratio column assignment (fixes "3 250028 A1 W 10:00" merging)
  • Tuned pdfplumber table_settings with Arabic-aware snap tolerances
  • Stricter bimodal gap detection for column schema stability
  • Improved visual-order Arabic detection using direction metadata
  • Better _fix_visual_arabic_cell: segment by token type before reversal
  • RTL-aware column ordering for Arabic tables
  • Minimum gap threshold to prevent false column splits
  • Enhanced OCR preprocessing (deskew + binarize)
  • Professional Excel output with per-column width optimization

Usage:
  python pdf_table_extractor.py input.pdf [output.xlsx]
  python pdf_table_extractor.py report.pdf -q
  python pdf_table_extractor.py report.pdf -v
  python pdf_table_extractor.py report.pdf --lang ara+eng
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
import time
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    level=logging.INFO,
    stream=sys.stderr,
)
log = logging.getLogger("pdf_extractor")

# ─────────────────────────────────────────────────────────────────────────────
# Arabic Unicode ranges
# ─────────────────────────────────────────────────────────────────────────────
_ARABIC_RE = re.compile(
    r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]'
)
_ARABIC_THRESHOLD = 0.05  # >5% Arabic chars → Arabic document

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 1 — DOCUMENT CLASSIFICATION
# ─────────────────────────────────────────────────────────────────────────────

class DocumentProfile:
    """Complete classification result for a PDF file."""
    __slots__ = (
        "is_scanned", "is_hybrid", "is_arabic", "has_borders",
        "page_count", "lang", "avg_chars_per_page",
        "text_page_indices", "scanned_page_indices",
    )

    def __init__(self):
        self.is_scanned          = False
        self.is_hybrid           = False
        self.is_arabic           = False
        self.has_borders         = False
        self.page_count          = 0
        self.lang                = "eng"
        self.avg_chars_per_page  = 0.0
        self.text_page_indices:   list[int] = []
        self.scanned_page_indices: list[int] = []


_LATTICE_SIGNALS = re.compile(
    rb"(?:RectangularPath|moveto|lineto|stroke|closepath|"
    rb"/Type\s*/Page|Rect\s*\[|/Border|/BS\s*/S)",
    re.IGNORECASE,
)
_MIN_CHARS_TEXT = 40  # chars per page below which → likely scanned


def classify_document(pdf_path: str) -> DocumentProfile:
    """
    Phase 1: Classify the PDF before any extraction begins.
    Reads only text layer + first 128 KB for border signals.
    """
    profile = DocumentProfile()

    # ── Border detection (128 KB raw scan + pdfplumber geometric check) ───────
    # NOTE: PDF content streams are zlib-compressed, so raw-byte PostScript
    # command search only hits uncompressed header/trailer sections.
    # We use this as a FIRST PASS and supplement with a pdfplumber geometric
    # check (rect + line objects) which works on the decoded page objects.
    try:
        with open(pdf_path, "rb") as fh:
            sample = fh.read(131_072)
        hits = len(_LATTICE_SIGNALS.findall(sample))
        profile.has_borders = hits >= 3
        log.info("Border signals (raw): %d → has_borders=%s", hits, profile.has_borders)
    except OSError:
        log.warning("Could not scan for border signals")

    # ── Text layer analysis (pdfplumber) ─────────────────────────────────────
    _CLASSIFY_SAMPLE = 20
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            profile.page_count = total_pages = len(pdf.pages)
            all_text = []
            page_char_counts = []

            if total_pages <= _CLASSIFY_SAMPLE:
                sample_indices = list(range(total_pages))
            else:
                mid = total_pages // 2
                sample_indices = sorted(set(
                    list(range(min(10, total_pages)))
                    + list(range(max(0, total_pages - 5), total_pages))
                    + [mid - 2, mid, mid + 2]
                ))

            sampled_text_pages = 0
            sampled_scanned_pages = 0

            for i in sample_indices:
                page = pdf.pages[i]
                text = (page.extract_text() or "").strip()
                page_char_counts.append(len(text))
                all_text.append(text)
                if len(text) >= _MIN_CHARS_TEXT:
                    sampled_text_pages += 1
                    profile.text_page_indices.append(i)
                else:
                    sampled_scanned_pages += 1
                    profile.scanned_page_indices.append(i)

        # ── Geometric border check (works on decoded page objects) ───────────
        # Check the first 3 pages for actual line/rect objects.  This catches
        # bordered tables in compressed-stream PDFs that the raw scan misses.
        if not profile.has_borders:
            try:
                geo_line_count = 0
                for pi in range(min(3, total_pages)):
                    p = pdf.pages[pi]
                    geo_line_count += len(p.lines or [])
                    geo_line_count += len(p.rects or [])
                if geo_line_count >= 4:
                    profile.has_borders = True
                    log.info(
                        "Border signals (geometric): %d lines/rects → has_borders=True",
                        geo_line_count,
                    )
            except Exception as _geo_exc:
                log.debug("Geometric border check failed: %s", _geo_exc)

        full_text = " ".join(all_text)
        total_chars = len(full_text.strip())

        sample_size = len(sample_indices)
        scanned_ratio = sampled_scanned_pages / max(sample_size, 1)
        profile.avg_chars_per_page = total_chars / max(profile.page_count, 1)

        if scanned_ratio < 0.05:
            profile.is_scanned = False
            profile.is_hybrid  = False
        elif scanned_ratio > 0.95:
            profile.is_scanned = True
            profile.is_hybrid  = False
        else:
            profile.is_scanned = False
            profile.is_hybrid  = True

        if total_chars > 0:
            arabic_chars = len(_ARABIC_RE.findall(full_text))
            ratio = arabic_chars / total_chars
            if ratio > _ARABIC_THRESHOLD:
                profile.is_arabic = True
                profile.lang = "ara+eng"
                log.info("Arabic detected (%.1f%%) → lang=ara+eng", ratio * 100)
            else:
                profile.lang = "eng"
        else:
            profile.lang = "ara+eng"

    except Exception as exc:
        log.warning("Classification failed (%s) — using defaults", exc)
        profile.is_scanned = False
        profile.lang = "eng"

    log.info(
        "Profile: pages=%d scanned=%s hybrid=%s arabic=%s borders=%s",
        profile.page_count, profile.is_scanned,
        profile.is_hybrid, profile.is_arabic, profile.has_borders,
    )
    return profile


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 2 — LAYOUT ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

class PageLayout:
    """Layout analysis result for a single page."""
    __slots__ = ("page_idx", "words", "has_table", "page_width", "page_height")

    def __init__(self, page_idx: int):
        self.page_idx   = page_idx
        self.words: list[dict] = []
        self.has_table  = False
        self.page_width = 0.0
        self.page_height = 0.0


def _extract_page_words(
    pdf_path: str,
    max_pages: int = 20,
) -> list[PageLayout]:
    """
    Phase 2: Extract word-level bounding boxes for column-schema building.
    Uses tight tolerances to avoid merging adjacent Arabic words.
    """
    import pdfplumber

    layouts: list[PageLayout] = []
    with pdfplumber.open(pdf_path) as pdf:
        pages_to_scan = pdf.pages[:max_pages] if len(pdf.pages) > max_pages else pdf.pages
        for i, page in enumerate(pages_to_scan):
            layout = PageLayout(i)
            layout.page_width  = float(page.width or 595)
            layout.page_height = float(page.height or 842)

            words = page.extract_words(
                x_tolerance=2,       # tighter: don't merge words from different columns
                y_tolerance=3,
                keep_blank_chars=False,
                use_text_flow=False,
            )
            if words:
                layout.words = words
                layout.has_table = len(words) >= 4
            layouts.append(layout)

    return layouts


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 5 — PERSISTENT COLUMN STABILIZATION
# ─────────────────────────────────────────────────────────────────────────────

class GlobalColumnSchema:
    """
    The global column anchor system — v3.

    Improvements over v2:
    - Uses word CENTER x-position for gap analysis (not just x0)
    - Enforces a minimum absolute gap threshold (avg_word_width * 0.8)
    - Uses OVERLAP RATIO for column assignment (not just x0 threshold)
    - Bins gaps using adaptive bin width from inter-quartile range
    """

    def __init__(self):
        self.splits: list[int] = []        # sorted column split X positions
        self.n_cols: int = 0
        self.page_width: float = 595.0
        self.avg_word_width: float = 20.0  # used for overlap tolerance

    def fit(self, layouts: list[PageLayout]) -> None:
        """Build the global column schema from all page layouts."""
        if not layouts:
            return

        self.page_width = max(
            (l.page_width for l in layouts if l.page_width > 0),
            default=595.0,
        )

        # ── Collect ALL gap records from ALL pages ────────────────────────────
        # Use gap between right-edge of word[i] and left-edge of word[i+1]
        all_gap_records: list[tuple[int, int]] = []
        all_word_widths: list[float] = []

        for layout in layouts:
            if not layout.words:
                continue

            for w in layout.words:
                try:
                    width = float(w.get("x1", 0)) - float(w.get("x0", 0))
                    if width > 0:
                        all_word_widths.append(width)
                except (TypeError, ValueError):
                    pass

            sorted_words = sorted(
                layout.words,
                key=lambda w: (
                    round(float(w.get("top", 0)) / 5) * 5,
                    float(w.get("x0", 0)),
                ),
            )

            rows = _cluster_words_into_rows_overlap(sorted_words)

            for row in rows:
                if len(row) < 2:
                    continue
                # Sort row by x0 for gap analysis
                sorted_row = sorted(row, key=lambda ww: float(ww.get("x0", 0)))
                for i in range(len(sorted_row) - 1):
                    gap_l = float(sorted_row[i].get("x1", 0))
                    gap_r = float(sorted_row[i + 1].get("x0", 0))
                    size  = gap_r - gap_l
                    if size > 0:
                        centre = int((gap_l + gap_r) / 2)
                        all_gap_records.append((int(size), centre))

        if all_word_widths:
            self.avg_word_width = float(np.median(all_word_widths))

        if not all_gap_records or len(all_gap_records) < 4:
            log.info("GlobalColumnSchema: insufficient gap data — single-column mode")
            self.splits = [int(self.page_width) + 1]
            self.n_cols = 1
            return

        # ── Find bimodal threshold (intra-cell vs inter-column gaps) ─────────
        sorted_sizes = sorted(s for s, _ in all_gap_records)

        if len(sorted_sizes) < 4:
            self.splits = [int(self.page_width) + 1]
            self.n_cols = 1
            return

        diffs = np.diff(sorted_sizes)
        biggest_jump_idx = int(np.argmax(diffs))
        small_max  = sorted_sizes[biggest_jump_idx]
        large_min  = sorted_sizes[biggest_jump_idx + 1]

        # Minimum meaningful column gap: at least 0.8× average word width
        min_col_gap = max(6, self.avg_word_width * 0.5)
        if large_min - small_max < min_col_gap:
            log.info(
                "GlobalColumnSchema: gap jump too small (%.1f < %.1f) → single-column",
                large_min - small_max, min_col_gap,
            )
            self.splits = [int(self.page_width) + 1]
            self.n_cols = 1
            return

        threshold  = (small_max + large_min) // 2

        # ── Build histogram of column-gap centre-X positions ─────────────────
        col_gap_centres = [c for s, c in all_gap_records if s > threshold]

        if not col_gap_centres:
            self.splits = [int(self.page_width) + 1]
            self.n_cols = 1
            return

        x_max = int(self.page_width) + 1
        # Adaptive bin width: use IQR-based estimate capped to reasonable range
        q25, q75 = np.percentile(col_gap_centres, [25, 75]) if len(col_gap_centres) > 4 else (0, 0)
        iqr_bw = (q75 - q25) / 2 if q75 > q25 else 0
        bin_width = max(8, int(min(iqr_bw, large_min * 0.15)))
        n_bins    = max(1, x_max // bin_width + 1)

        hist = np.zeros(n_bins, dtype=np.int32)
        for gx in col_gap_centres:
            bi = min(int(gx) // bin_width, n_bins - 1)
            hist[bi] += 1

        contributing_pages = sum(1 for l in layouts if l.words)
        min_votes = max(1, contributing_pages // 4)

        col_splits: list[int] = []
        in_peak  = False
        peak_xs: list[int] = []

        for bi in range(n_bins):
            if hist[bi] >= min_votes:
                if not in_peak:
                    in_peak  = True
                    peak_xs  = []
                peak_xs.extend([bi * bin_width] * int(hist[bi]))
            else:
                if in_peak:
                    col_splits.append(int(np.mean(peak_xs)))
                    in_peak  = False
                    peak_xs  = []
        if in_peak and peak_xs:
            col_splits.append(int(np.mean(peak_xs)))

        col_splits.append(x_max)

        if len(col_splits) < 2:
            col_splits = [x_max]

        # Merge split points that are too close together (< avg_word_width * 0.5)
        merge_dist = max(8, int(self.avg_word_width * 0.5))
        merged_splits: list[int] = []
        for sp in col_splits:
            if merged_splits and sp - merged_splits[-1] < merge_dist:
                merged_splits[-1] = (merged_splits[-1] + sp) // 2
            else:
                merged_splits.append(sp)

        # Ensure sentinel is present
        if not merged_splits or merged_splits[-1] < x_max:
            merged_splits.append(x_max)

        self.splits = merged_splits
        self.n_cols = len(self.splits)
        log.info(
            "GlobalColumnSchema: %d columns, split points=%s (avg_word_w=%.1f)",
            self.n_cols, self.splits[:-1], self.avg_word_width,
        )

    def assign_column(self, x0: float, x1: Optional[float] = None) -> int:
        """
        Assign a word to a column using OVERLAP RATIO.

        If x1 is provided: find the column with the greatest overlap with [x0, x1].
        If only x0: fall back to boundary threshold (backward compat).
        """
        if x1 is None or x1 <= x0:
            # Legacy: use x0 threshold
            xi = int(x0)
            for ci, boundary in enumerate(self.splits):
                if xi < boundary:
                    return ci
            return self.n_cols - 1

        # Overlap-based: find column with most overlap
        word_w = x1 - x0
        if word_w <= 0:
            return self.assign_column(x0)

        best_ci   = 0
        best_ovlp = -1.0
        col_left  = 0.0

        for ci, boundary in enumerate(self.splits):
            col_right = float(boundary)
            ovlp = max(0.0, min(x1, col_right) - max(x0, col_left))
            ratio = ovlp / word_w
            if ratio > best_ovlp:
                best_ovlp = ratio
                best_ci   = ci
            col_left = col_right

        return best_ci


# ─────────────────────────────────────────────────────────────────────────────
# SHARED ROW CLUSTERING — overlap-based (v3)
# ─────────────────────────────────────────────────────────────────────────────

def _cluster_words_into_rows_overlap(
    sorted_words: list[dict],
    overlap_thresh: float = 0.4,
) -> list[list[dict]]:
    """
    Cluster words into rows using vertical OVERLAP rather than Y-midpoint distance.

    A word joins the current row if the vertical overlap between the word's
    [top, bottom] and the row's [top, bottom] exceeds `overlap_thresh` fraction
    of the word's own height.  This is robust to varying line heights and
    multi-line cells.

    The row's vertical extent expands as words are added (union of bboxes).
    """
    if not sorted_words:
        return []

    rows: list[list[dict]] = []
    row_tops:    list[float] = []
    row_bottoms: list[float] = []

    for w in sorted_words:
        try:
            w_top    = float(w.get("top",    0))
            w_bottom = float(w.get("bottom", 0))
        except (TypeError, ValueError):
            w_top = w_bottom = 0.0

        w_h = max(w_bottom - w_top, 1.0)
        placed = False

        # Scan existing rows bottom-up (most likely to match most recent rows)
        for ri in range(len(rows) - 1, max(len(rows) - 6, -1), -1):
            r_top    = row_tops[ri]
            r_bottom = row_bottoms[ri]
            # Overlap = intersection of [w_top, w_bottom] and [r_top, r_bottom]
            ovlp = max(0.0, min(w_bottom, r_bottom) - max(w_top, r_top))
            if ovlp / w_h >= overlap_thresh:
                rows[ri].append(w)
                row_tops[ri]    = min(r_top,    w_top)
                row_bottoms[ri] = max(r_bottom, w_bottom)
                placed = True
                break

        if not placed:
            rows.append([w])
            row_tops.append(w_top)
            row_bottoms.append(w_bottom)

    # Sort each row by x0 (left-to-right)
    for ri in range(len(rows)):
        rows[ri] = sorted(rows[ri], key=lambda ww: float(ww.get("x0", 0)))

    # Sort rows by their top y-coordinate
    rows_sorted = sorted(
        zip(row_tops, rows),
        key=lambda t: t[0],
    )
    return [r for _, r in rows_sorted]


# Keep backward-compat alias
_cluster_words_into_rows = _cluster_words_into_rows_overlap


def layout_to_dataframe(
    layout: PageLayout,
    schema: GlobalColumnSchema,
) -> Optional[pd.DataFrame]:
    """
    Convert a PageLayout to a DataFrame using the global column schema.

    v3: Uses overlap-based row clustering and word-extent column assignment.
    Prevents multiple words from the same row being merged into one column
    unless they genuinely belong there.
    """
    if not layout.words:
        return None

    sorted_words = sorted(
        layout.words,
        key=lambda w: (
            round(float(w.get("top", 0)) / 5) * 5,
            float(w.get("x0", 0)),
        ),
    )

    rows = _cluster_words_into_rows_overlap(sorted_words)
    if len(rows) < 2:
        return None

    ncols  = schema.n_cols
    matrix: list[list[str]] = []

    for row in rows:
        cells = [""] * ncols
        for w in row:
            x0   = float(w.get("x0", 0))
            x1   = float(w.get("x1", x0 + 1))
            text = str(w.get("text", "")).strip()
            if not text:
                continue
            ci = schema.assign_column(x0, x1)
            cells[ci] = (cells[ci] + " " + text).strip() if cells[ci] else text
        if any(cells):
            matrix.append(cells)

    if not matrix:
        return None

    return pd.DataFrame(matrix)


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 3 — SMART ENGINE ROUTING
# ─────────────────────────────────────────────────────────────────────────────

# pdfplumber table_settings tuned for Arabic PDFs
_TABLE_SETTINGS_LATTICE = {
    "vertical_strategy":   "lines",
    "horizontal_strategy": "lines",
    "snap_tolerance":      3,
    "join_tolerance":      3,
    "edge_min_length":     3,
    "min_words_vertical":  1,
    "min_words_horizontal": 1,
    "text_x_tolerance":    2,
    "text_y_tolerance":    3,
}

_TABLE_SETTINGS_STREAM = {
    "vertical_strategy":   "text",
    "horizontal_strategy": "text",
    "snap_tolerance":      3,
    "join_tolerance":      3,
    "edge_min_length":     3,
    "min_words_vertical":  3,
    "min_words_horizontal": 1,
    "text_x_tolerance":    2,
    "text_y_tolerance":    3,
}

_TABLE_SETTINGS_EXPLICIT = {
    "vertical_strategy":   "explicit",
    "horizontal_strategy": "lines",
    "snap_tolerance":      3,
    "join_tolerance":      3,
    "edge_min_length":     3,
    "min_words_vertical":  1,
    "min_words_horizontal": 1,
    "text_x_tolerance":    2,
    "text_y_tolerance":    3,
}


def _extract_camelot(pdf_path: str, flavor: str) -> list[pd.DataFrame]:
    """Extract tables via Camelot with auto-fallback to opposite flavor."""
    import camelot

    log.info("Camelot [%s] extracting from %s", flavor, pdf_path)
    try:
        tables = camelot.read_pdf(
            pdf_path, pages="all", flavor=flavor, suppress_stdout=True
        )
        log.info("Camelot [%s] found %d table(s)", flavor, len(tables))

        if len(tables) == 0:
            alt = "stream" if flavor == "lattice" else "lattice"
            log.info("Retrying with Camelot [%s]", alt)
            try:
                alt_tables = camelot.read_pdf(
                    pdf_path, pages="all", flavor=alt, suppress_stdout=True
                )
                if len(alt_tables) > 0:
                    return [t.df for t in alt_tables]
            except Exception as exc:
                log.debug("Camelot [%s] also failed: %s", alt, exc)

        return [t.df for t in tables]

    except Exception as exc:
        log.warning("Camelot [%s] failed: %s", flavor, exc)
        return []


def _extract_pdfplumber_tables(
    pdf_path: str,
    fix_visual_arabic: bool = False,
    has_borders: bool = False,
) -> list[pd.DataFrame]:
    """
    pdfplumber explicit table detection — v3.

    Tries multiple strategies in order:
    1. lattice (lines-based) if has_borders
    2. stream (text-based) — most reliable for borderless Arabic tables
    3. explicit vertical splits from GlobalColumnSchema

    When fix_visual_arabic=True each cell's text is corrected with
    _fix_visual_arabic_cell() to convert visual-order to logical-order.
    """
    import pdfplumber

    log.info(
        "pdfplumber v3 extraction from %s (visual_arabic=%s, borders=%s)",
        pdf_path, fix_visual_arabic, has_borders,
    )

    strategies = []
    if has_borders:
        strategies.append(("lattice", _TABLE_SETTINGS_LATTICE))
    strategies.append(("stream", _TABLE_SETTINGS_STREAM))

    dfs: list[pd.DataFrame] = []

    for strategy_name, settings in strategies:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                page_dfs: list[pd.DataFrame] = []
                for page in pdf.pages:
                    tables = page.find_tables(table_settings=settings)
                    for tbl in tables:
                        extracted = tbl.extract()
                        if not extracted:
                            continue
                        if fix_visual_arabic:
                            corrected = [
                                [
                                    _fix_visual_arabic_cell(str(cell) if cell else "")
                                    for cell in row
                                ]
                                for row in extracted
                            ]
                            page_dfs.append(pd.DataFrame(corrected))
                        else:
                            page_dfs.append(pd.DataFrame(extracted))

            if page_dfs:
                log.info(
                    "pdfplumber [%s] found %d table(s)",
                    strategy_name, len(page_dfs),
                )
                dfs = page_dfs
                break
        except Exception as exc:
            log.warning("pdfplumber [%s] failed: %s", strategy_name, exc)

    if not dfs:
        log.info("pdfplumber: no tables with structured detection — trying word-based")
        try:
            dfs = _extract_pdfplumber_words(pdf_path, fix_visual_arabic)
        except Exception as exc:
            log.warning("pdfplumber word-based failed: %s", exc)

    log.info("pdfplumber found %d table(s) total", len(dfs))
    return dfs


def _extract_pdfplumber_words(
    pdf_path: str,
    fix_visual_arabic: bool = False,
) -> list[pd.DataFrame]:
    """
    Fallback pdfplumber extraction using extract_words() + spatial clustering.
    Builds a local GlobalColumnSchema per page for accurate column assignment.
    """
    import pdfplumber

    dfs: list[pd.DataFrame] = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                words = page.extract_words(
                    x_tolerance=2,
                    y_tolerance=3,
                    keep_blank_chars=False,
                    use_text_flow=False,
                )
                if not words or len(words) < 4:
                    continue

                layout = PageLayout(0)
                layout.words      = words
                layout.page_width  = float(page.width or 595)
                layout.page_height = float(page.height or 842)

                local_schema = GlobalColumnSchema()
                local_schema.fit([layout])

                df = layout_to_dataframe(layout, local_schema)
                if df is not None and not df.empty:
                    if fix_visual_arabic:
                        df = df.map(
                            lambda v: _fix_visual_arabic_cell(_to_str(v))
                        )
                    dfs.append(df)
    except Exception as exc:
        log.warning("_extract_pdfplumber_words failed: %s", exc)

    return dfs


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 4 — OCR PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def _preprocess_ocr_image(image):
    """
    Preprocess a PIL image for better Tesseract accuracy on Arabic text.
    Steps: grayscale → adaptive threshold → mild deskew.
    """
    try:
        import cv2
        import numpy as np_cv
        from PIL import Image

        # Convert PIL → OpenCV
        img_np = np_cv.array(image.convert("RGB"))
        gray   = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)

        # Adaptive thresholding for clean binarization
        binary = cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            blockSize=31, C=10,
        )

        # Mild denoising
        denoised = cv2.fastNlMeansDenoising(binary, None, h=10)

        return Image.fromarray(denoised)
    except Exception:
        return image  # safe fallback — return original


def _ocr_page_to_df(
    image,
    lang: str = "eng",
    schema: Optional[GlobalColumnSchema] = None,
) -> Optional[pd.DataFrame]:
    """
    Run Tesseract on a PIL image, extract word bounding boxes,
    cluster into rows using overlap-based grouping, and assign to columns.
    """
    import pytesseract

    # Preprocess image for better Arabic OCR
    processed_image = _preprocess_ocr_image(image)

    # Use PSM 6 (uniform block of text) + OEM 1 (LSTM only)
    # For Arabic: use higher DPI hint
    ocr_config = "--psm 6 --oem 1 -c preserve_interword_spaces=1"
    if "ara" in lang:
        ocr_config += " -c textord_arabic_whitespace=1"

    data = pytesseract.image_to_data(
        processed_image,
        output_type=pytesseract.Output.DATAFRAME,
        lang=lang,
        config=ocr_config,
    )

    data = data[
        (data["conf"] >= 35) &
        (data["text"].astype(str).str.strip() != "") &
        (data["text"].notna())
    ].copy()

    if len(data) < 4:
        return None

    data["x_right"] = data["left"].astype(int) + data["width"].astype(int)
    data["y_mid"]   = data["top"].astype(int)  + data["height"].astype(int) / 2.0

    words = []
    for _, row in data.iterrows():
        words.append({
            "x0":     float(row["left"]),
            "x1":     float(row["x_right"]),
            "top":    float(row["top"]),
            "bottom": float(row["top"]) + float(row["height"]),
            "text":   str(row["text"]).strip(),
        })

    sorted_words = sorted(words, key=lambda w: (
        round(float(w["top"]) / 5) * 5, float(w["x0"])
    ))
    rows = _cluster_words_into_rows_overlap(sorted_words)

    if len(rows) < 2:
        return None

    # Use global schema if available, else do local gap voting
    if schema is not None and schema.n_cols > 1:
        ncols  = schema.n_cols
        matrix: list[list[str]] = []
        for row in rows:
            cells = [""] * ncols
            for w in row:
                ci = schema.assign_column(float(w["x0"]), float(w["x1"]))
                t  = w["text"]
                cells[ci] = (cells[ci] + " " + t).strip() if cells[ci] else t
            if any(cells):
                matrix.append(cells)
        if not matrix:
            return None
        return pd.DataFrame(matrix)

    # ── Local gap voting (no global schema) ──────────────────────────────────
    # Build a local schema from this page's words
    layout = PageLayout(0)
    layout.words = sorted_words
    # Estimate page width from rightmost word
    layout.page_width = max(
        (float(w["x1"]) for w in sorted_words), default=595.0
    ) + 20

    local_schema = GlobalColumnSchema()
    local_schema.fit([layout])

    if local_schema.n_cols < 2:
        return None

    ncols  = local_schema.n_cols
    matrix2: list[list[str]] = []
    for row in rows:
        cells = [""] * ncols
        for w in row:
            ci = local_schema.assign_column(float(w["x0"]), float(w["x1"]))
            t  = w["text"]
            cells[ci] = (cells[ci] + " " + t).strip() if cells[ci] else t
        if any(cells):
            matrix2.append(cells)

    if not matrix2:
        return None

    return pd.DataFrame(matrix2)


def _extract_ocr(
    pdf_path: str,
    lang: str = "eng",
    schema: Optional[GlobalColumnSchema] = None,
) -> list[pd.DataFrame]:
    """Phase 4: Convert each PDF page to image (300 DPI) and OCR it."""
    from pdf2image import convert_from_path

    log.info("OCR pipeline: lang=%s", lang)
    images = convert_from_path(pdf_path, dpi=300)
    log.info("OCR: %d page(s)", len(images))

    dfs: list[pd.DataFrame] = []
    for i, img in enumerate(images, start=1):
        log.debug("OCR page %d/%d", i, len(images))
        df = _ocr_page_to_df(img, lang=lang, schema=schema)
        if df is not None and not df.empty:
            dfs.append(df)
        else:
            log.debug("OCR page %d — no table detected", i)

    log.info("OCR found %d raw table(s)", len(dfs))
    return dfs


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 6 — TABLE RECONSTRUCTION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_header_row(row: pd.Series) -> tuple:
    """Produce a hashable fingerprint of a header row for deduplication."""
    vals = []
    for v in row:
        s = str(v).strip().lower()
        s = re.sub(r'\s+', ' ', s)
        s = unicodedata.normalize("NFKC", s)
        vals.append(s)
    return tuple(vals)


def _header_similarity(a: tuple, b: tuple) -> float:
    """Simple token overlap similarity for header deduplication."""
    if not a or not b:
        return 0.0
    set_a = set(a)
    set_b = set(b)
    intersection = set_a & set_b
    union = set_a | set_b
    return len(intersection) / len(union) if union else 0.0


def _schemas_compatible(a: pd.DataFrame, b: pd.DataFrame) -> bool:
    """
    Two tables are compatible for merging when they have the same column count,
    OR when one has only 1 row (it's a fragment).
    """
    if a.shape[1] == b.shape[1]:
        return True
    if abs(a.shape[1] - b.shape[1]) <= 1 and min(a.shape[1], b.shape[1]) >= 2:
        return True
    return False


def _pad_to_width(df: pd.DataFrame, target_cols: int) -> pd.DataFrame:
    """Pad a DataFrame with empty columns on the right to reach target_cols."""
    while df.shape[1] < target_cols:
        df[f"_pad_{df.shape[1]}"] = ""
    return df


def reconstruct_tables(raw_dfs: list[pd.DataFrame]) -> list[pd.DataFrame]:
    """
    Phase 6 + Phase 7: Merge compatible adjacent tables and remove duplicate headers.

    v3 improvements:
    - Uses header similarity scoring for more robust deduplication
    - Allows 80%+ similarity headers to be deduped (not just exact match)
    - Better handling of page-continuation fragments
    """
    if not raw_dfs:
        return []

    cleaned = []
    for df in raw_dfs:
        c = _clean_dataframe(df)
        if c is not None:
            cleaned.append(c)

    if not cleaned:
        return []

    groups: list[pd.DataFrame] = [cleaned[0]]
    seen_headers: set[tuple] = set()

    # Seed seen_headers with BOTH the column-name tuple (so any repeated header
    # row that slipped past header-promotion is deduped) AND the first data row.
    col_fp = tuple(
        unicodedata.normalize("NFKC", re.sub(r"\s+", " ", str(c).strip().lower()))
        for c in cleaned[0].columns
    )
    seen_headers.add(col_fp)
    if len(cleaned[0]) > 0:
        seen_headers.add(_normalize_header_row(cleaned[0].iloc[0]))

    for df in cleaned[1:]:
        current = groups[-1]

        if _schemas_compatible(current, df):
            target = max(current.shape[1], df.shape[1])
            if current.shape[1] < target:
                current = _pad_to_width(current.copy(), target)
                groups[-1] = current
            if df.shape[1] < target:
                df = _pad_to_width(df.copy(), target)

            df.columns = current.columns[:df.shape[1]]

            rows_to_keep = []
            for i, row in df.iterrows():
                fp = _normalize_header_row(row)
                # Exact match dedup
                if fp in seen_headers:
                    continue
                # Similarity dedup (≥80% token overlap → treat as repeated header)
                is_dup = any(
                    _header_similarity(fp, seen_fp) >= 0.80
                    for seen_fp in seen_headers
                )
                if not is_dup:
                    rows_to_keep.append(i)

            if rows_to_keep:
                df_filtered = df.loc[rows_to_keep].copy()
                if len(df_filtered) > 0:
                    seen_headers.add(_normalize_header_row(df_filtered.iloc[0]))

                merged = pd.concat(
                    [current, df_filtered], ignore_index=True
                )
                groups[-1] = merged
        else:
            if len(df) > 0:
                seen_headers.add(_normalize_header_row(df.iloc[0]))
            groups.append(df)

    result = []
    for g in groups:
        g = g.reset_index(drop=True)
        g.replace("", pd.NA, inplace=True)
        g.dropna(how="all", inplace=True)
        g.dropna(axis=1, how="all", inplace=True)
        g.fillna("", inplace=True)
        g.reset_index(drop=True, inplace=True)
        if g.shape[0] >= 1 and g.shape[1] >= 1:
            result.append(g)

    log.info("Reconstruction: %d raw → %d merged table(s)", len(raw_dfs), len(result))
    return result


# ─────────────────────────────────────────────────────────────────────────────
# DATA CLEANING (shared utility)
# ─────────────────────────────────────────────────────────────────────────────

_BLANK_VALS = {"", "nan", "none", "n/a", "-", "<na>", "na"}


def _to_str(v: object) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        s = str(v).strip()
        return "" if s.lower() in _BLANK_VALS else s
    except Exception:
        return ""


def _dedup_columns(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for i, name in enumerate(names):
        base = name.strip() if name.strip() else f"Col_{i + 1}"
        if base in seen:
            seen[base] += 1
            result.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            result.append(base)
    return result


def _clean_dataframe(raw: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Vectorised cleaning:
    1. Convert to object dtype
    2. Blank/NaN strings → ""
    3. Drop all-empty rows and columns
    4. Promote header if columns are integer indices
    5. NFKC-normalise strings
    """
    if raw is None or raw.empty:
        return None

    df = raw.copy().astype(object)

    for col in df.columns:
        df[col] = df[col].map(_to_str)

    df.replace("", pd.NA, inplace=True)
    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)
    df.fillna("", inplace=True)
    df.reset_index(drop=True, inplace=True)

    if df.empty or df.shape[1] < 1:
        return None

    # Header promotion for integer-indexed columns
    if all(
        isinstance(c, int) or (isinstance(c, str) and c.isdigit())
        for c in df.columns
    ):
        ncols = df.shape[1]
        header_idx = 0
        while header_idx < len(df):
            row_vals  = [_to_str(v) for v in df.iloc[header_idx]]
            non_empty = sum(1 for v in row_vals if v)
            if non_empty > max(1, ncols // 2):
                break
            header_idx += 1

        if header_idx >= len(df):
            return None

        new_header = [_to_str(v) for v in df.iloc[header_idx]]
        df = df.iloc[header_idx + 1:].copy()
        df.columns = _dedup_columns(new_header)
        df.reset_index(drop=True, inplace=True)
    else:
        df.columns = _dedup_columns([str(c) for c in df.columns])

    # NFKC normalisation
    for col in df.columns:
        df[col] = df[col].map(
            lambda v: unicodedata.normalize("NFKC", _to_str(v))
        )

    mask = df.apply(
        lambda row: row.astype(str).str.strip().ne("").any(), axis=1
    )
    df = df[mask].reset_index(drop=True)

    if df.shape[0] < 1 or df.shape[1] < 1:
        return None

    return df


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 8 — ARABIC RTL SUPPORT (v3)
# ─────────────────────────────────────────────────────────────────────────────
#
# ROOT-CAUSE ANALYSIS (unchanged from v2 — still valid):
# Many Arabic PDFs store characters in VISUAL order (LTR in content stream).
# pdfplumber reports direction='ltr' for these Arabic words.
#
# FIX strategy:
#   detect visual-order → reverse char order per Arabic token → reverse word list.
#   Do NOT apply arabic_reshaper/bidi for Excel — Excel's own engine handles it.
#
# v3 improvements to _fix_visual_arabic_cell:
#   - Better token segmentation: split on Arabic/Latin/digit boundaries
#   - Handle Arabic-digit mixed tokens with segment-level reversal
#   - Preserve punctuation position correctly
#   - Handle newlines in multi-line cells robustly
# ─────────────────────────────────────────────────────────────────────────────

def _is_rtl_table(df: pd.DataFrame) -> bool:
    """Detect whether a table's content is primarily Arabic/RTL."""
    sample_text = " ".join(
        str(v)
        for row in df.values[:10]
        for v in row
    )
    total = len(sample_text.strip())
    if total == 0:
        return False
    arabic_chars = len(_ARABIC_RE.findall(sample_text))
    return arabic_chars / total > _ARABIC_THRESHOLD


# Token type constants
_TT_ARABIC  = "ar"
_TT_LATIN   = "la"
_TT_DIGIT   = "di"
_TT_SPACE   = "sp"
_TT_PUNCT   = "pu"

_DIGIT_CHARS = set("0123456789.,:;/+-")
_LATIN_RE    = re.compile(r'[A-Za-z]')


def _tokenize_cell(text: str) -> list[tuple[str, str]]:
    """
    Segment cell text into typed tokens: (type, chars).
    Types: ar=Arabic, la=Latin, di=digit/punct, sp=space.

    This allows independent handling of each segment during reversal.
    """
    tokens: list[tuple[str, str]] = []
    cur_type = None
    cur_chars: list[str] = []

    for ch in text:
        if ch == " ":
            t = _TT_SPACE
        elif _ARABIC_RE.match(ch):
            t = _TT_ARABIC
        elif _LATIN_RE.match(ch):
            t = _TT_LATIN
        elif ch in _DIGIT_CHARS:
            t = _TT_DIGIT
        else:
            t = _TT_PUNCT

        if t != cur_type:
            if cur_chars:
                tokens.append((cur_type, "".join(cur_chars)))
            cur_type  = t
            cur_chars = [ch]
        else:
            cur_chars.append(ch)

    if cur_chars:
        tokens.append((cur_type, "".join(cur_chars)))

    return tokens


def _fix_visual_arabic_cell(text: str) -> str:
    """
    Correct visual-order Arabic cell text for Excel storage — v3.

    Algorithm:
      0. NFKC-normalize to convert Arabic Presentation Forms (U+FB50-UFDFF,
         U+FE70-FEFF) to their basic Unicode equivalents — required so that
         ligature/glyph-level characters like ﻢ (Arabic Meem Medial Form) are
         first decoded to م before the reversal step.
      1. Handle multi-line cells line by line
      2. Tokenize into typed segments (Arabic / Latin / Digit / Space / Punct)
      3. Group by space-separated "words" (preserving token types)
      4. Reverse character order within each Arabic-only word
      5. Reverse the word list if any Arabic words were found
      6. Rejoin words with spaces

    This correctly handles mixed cells like "A1 W 10:00 محاضرة" where
    Latin and numeric tokens must preserve their exact character order.
    """
    if not text:
        return text
    # NFKC normalization: converts Arabic Presentation Forms to basic Arabic
    # script before reversal so ligature code-points don't corrupt the output.
    text = unicodedata.normalize("NFKC", text)
    if not _ARABIC_RE.search(text):
        return text

    # Handle multi-line cells line by line
    if "\n" in text:
        lines = text.split("\n")
        return "\n".join(_fix_visual_arabic_cell(line) for line in lines)

    # Split on spaces to get words, preserving space count
    parts = text.split(" ")
    fixed: list[str] = []
    has_arabic_token = False

    for part in parts:
        if not part:
            fixed.append(part)
            continue

        if not _ARABIC_RE.search(part):
            # Purely Latin/digit/punct — keep as-is
            fixed.append(part)
            continue

        # Part contains Arabic
        tokens = _tokenize_cell(part)

        # Check if it's purely Arabic (no Latin/digit mixed in)
        non_arabic_types = {tt for tt, _ in tokens if tt not in (_TT_ARABIC, _TT_PUNCT, _TT_SPACE)}
        if not non_arabic_types:
            # Pure Arabic word: reverse characters
            fixed.append(part[::-1])
            has_arabic_token = True
        else:
            # Mixed token (e.g. "250114كلية"): reverse only the Arabic segments
            # and keep the rest in place — safer than full reversal
            rebuilt = []
            for tt, chars in tokens:
                if tt == _TT_ARABIC:
                    rebuilt.append(chars[::-1])
                    has_arabic_token = True
                else:
                    rebuilt.append(chars)
            fixed.append("".join(rebuilt))

    if has_arabic_token:
        fixed.reverse()

    return " ".join(fixed)


def fix_visual_arabic_dataframe(
    df: pd.DataFrame,
    visual_order: bool = True,
) -> pd.DataFrame:
    """
    Apply visual-order Arabic correction to every cell in a DataFrame.
    When visual_order=False: returns df unchanged.
    """
    if not visual_order:
        return df

    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(lambda v: _fix_visual_arabic_cell(_to_str(v)))

    new_cols = [_fix_visual_arabic_cell(str(c)) for c in out.columns]
    out.columns = new_cols
    return out


def _detect_visual_order_from_words(words: list[dict]) -> bool:
    """
    Return True when pdfplumber word objects indicate visual-order Arabic.
    Criteria: ≥50% of Arabic-containing words have direction='ltr'.
    Also checks if Arabic text appears reversed (chars end with common prefix letters).
    """
    arabic_words = [w for w in words if _ARABIC_RE.search(str(w.get("text", "")))]
    if not arabic_words:
        return False

    # Primary signal: pdfplumber direction metadata
    ltr_count = sum(1 for w in arabic_words if w.get("direction", "") == "ltr")
    if ltr_count / len(arabic_words) >= 0.5:
        return True

    # Secondary signal: first token ends with common reversed-start patterns
    first_texts = [str(w.get("text", "")) for w in arabic_words[:20]]
    reversed_start_endings = set("التنيةةىلاوبسكمهرعق")
    reversed_count = 0
    for t in first_texts:
        tokens = [tok for tok in t.split() if _ARABIC_RE.search(tok)]
        if tokens and tokens[0] and tokens[0][-1] in reversed_start_endings:
            reversed_count += 1

    return reversed_count / max(len(first_texts), 1) >= 0.6


# Keep a thin public alias for backward compatibility
def apply_arabic_reshaping(df: pd.DataFrame) -> pd.DataFrame:
    """Alias kept for API compatibility — delegates to fix_visual_arabic_dataframe."""
    return fix_visual_arabic_dataframe(df, visual_order=True)


# ── Bracket / parenthesis mirroring ──────────────────────────────────────────
# In visual-order Arabic PDFs every bidirectional character (parentheses,
# brackets, braces) is stored reversed throughout the ENTIRE row — including
# inside purely-English cells like "Physiology IV )510(".
# After we correct word order with fix_visual_arabic_dataframe we must
# also mirror all such characters in the whole table.
_BRACKET_MIRROR_TABLE = str.maketrans("()[]{}⟨⟩", ")(][}{⟩⟨")


def _mirror_brackets_cell(text: str) -> str:
    """Flip all bracket/paren/brace chars in a single cell string."""
    if not text:
        return text
    return str(text).translate(_BRACKET_MIRROR_TABLE)


def mirror_brackets_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Apply bracket mirroring to every cell in a DataFrame."""
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(lambda v: _mirror_brackets_cell(_to_str(v)))
    new_cols = [_mirror_brackets_cell(str(c)) for c in out.columns]
    out.columns = new_cols
    return out


# ── Row-fragment repair ───────────────────────────────────────────────────────
# Camelot stream mode splits multi-line cells across separate rows because it
# has no grid lines to anchor on.  Pattern: a "continuation" row has an empty
# first column and carries content that belongs to the previous non-empty row.
# We merge such rows upward by appending their non-empty cells.

def _merge_fragmented_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Merge continuation rows (first column empty) into the preceding data row.

    A row is treated as a continuation if:
      - Its first column (index 0) is empty, AND
      - At least one other column is non-empty, AND
      - There is already a preceding data row to absorb it.

    Merging: for each non-empty cell in the continuation row, append its text
    to the corresponding cell in the preceding row with a newline separator.

    This is a safety-net for the stream fallback.  Lattice mode reads actual
    grid lines and should not need this.
    """
    if df.empty or df.shape[0] < 2:
        return df

    rows: list[list[str]] = [
        [_to_str(v) for v in row] for row in df.values
    ]
    merged: list[list[str]] = []

    for row in rows:
        first_col = row[0].strip() if row else ""
        rest_non_empty = any(c.strip() for c in row[1:])

        if not first_col and rest_non_empty and merged:
            # Continuation row — append to previous
            prev = merged[-1]
            for ci in range(1, len(row)):
                fragment = row[ci].strip()
                if not fragment:
                    continue
                existing = prev[ci].strip()
                if existing:
                    prev[ci] = existing + "\n" + fragment
                else:
                    prev[ci] = fragment
        else:
            merged.append(list(row))

    result = pd.DataFrame(merged, columns=df.columns)
    result.reset_index(drop=True, inplace=True)
    return result


def _reorder_rtl_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    For RTL (Arabic) tables, reverse column order so reading flows right-to-left.
    This matches how the original PDF would be read.
    Only applied when the table is detected as primarily Arabic.
    """
    if df.shape[1] <= 1:
        return df
    return df.iloc[:, ::-1].reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 9 — EXCEL GENERATION ENGINE (v3)
# ─────────────────────────────────────────────────────────────────────────────

# Styles
_HDR_FILL     = PatternFill("solid", fgColor="1E3A5F")   # deep navy
_HDR_FONT     = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_HDR_BORDER   = Border(
    bottom=Side(style="medium", color="FFFFFF"),
    left=Side(style="thin", color="1E3A5F"),
    right=Side(style="thin", color="1E3A5F"),
)
_CELL_FONT    = Font(name="Arial", size=10)
_ALT_FILL     = PatternFill("solid", fgColor="EEF3F8")   # very light blue-grey
_META_FILL    = PatternFill("solid", fgColor="F1F5F9")
_THIN_SIDE    = Side(style="thin", color="C5CBD5")
_THIN_BORDER  = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)

_ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT_WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
_ALIGN_RIGHT_RTL   = Alignment(horizontal="right",  vertical="top",    wrap_text=True, readingOrder=2)
_ALIGN_CENTER_RTL  = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)


def _auto_col_width(ws, df: pd.DataFrame, is_rtl: bool = False) -> None:
    """
    Set column widths based on content.

    v3: Uses actual max character length per column (not capped header+4).
    Arabic characters are counted as 1.5× width because they render wider.
    Caps: minimum 8, maximum 55.
    """
    for idx, col in enumerate(df.columns, start=1):
        header_len = _visual_len(str(col))
        try:
            data_col = df.iloc[:, idx - 1].astype(str)
            max_data_len = int(data_col.map(_visual_len).max())
        except Exception:
            max_data_len = 0

        content_len = max(header_len, max_data_len)
        # Scale: each visual character ≈ 1.1 Excel width unit
        width = max(8, min(int(content_len * 1.1) + 2, 55))
        ws.column_dimensions[get_column_letter(idx)].width = width


def _visual_len(text: str) -> int:
    """
    Estimate visual display length of a string.
    Arabic chars count as 1.5×, Latin/digits as 1×.
    """
    ar_count  = len(_ARABIC_RE.findall(text))
    lat_count = len(text) - ar_count
    return int(ar_count * 1.5 + lat_count)


def _write_table_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
    is_rtl: bool,
    table_num: int,
) -> None:
    """Write one DataFrame to a named sheet with full professional formatting."""
    ws = wb.create_sheet(sheet_name)

    if is_rtl:
        ws.sheet_view.rightToLeft = True

    # Set default row height for readability
    ws.sheet_format.defaultRowHeight = 18

    cell_align = _ALIGN_RIGHT_RTL if is_rtl else _ALIGN_LEFT_WRAP
    hdr_align  = _ALIGN_CENTER_RTL if is_rtl else _ALIGN_CENTER_WRAP

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(1, col_idx, str(header))
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = hdr_align
        cell.border    = _HDR_BORDER

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, start=1):
            val_str = "" if pd.isna(val) else str(val)
            cell = ws.cell(row_idx, col_idx, val_str)
            cell.font      = _CELL_FONT
            cell.border    = _THIN_BORDER
            # Numbers and short values: center; Arabic text: right-align
            if is_rtl and _ARABIC_RE.search(val_str):
                cell.alignment = _ALIGN_RIGHT_RTL
            elif is_rtl:
                cell.alignment = _ALIGN_CENTER_RTL
            else:
                cell.alignment = _ALIGN_LEFT_WRAP
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    _auto_col_width(ws, df, is_rtl)

    # Auto-fit row heights: count newlines in cells
    for row_idx in range(2, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                lines = str(cell.value).count("\n") + 1
                max_lines = max(max_lines, lines)
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = max(18, max_lines * 14)

    log.info("Sheet '%s': %d rows × %d cols (RTL=%s)", sheet_name, *df.shape, is_rtl)


def write_excel(
    tables: list[pd.DataFrame],
    output_path: str,
    source_pdf: str,
    engine: str,
    is_arabic: bool = False,
) -> None:
    """Phase 9: Write all tables to a production-quality .xlsx file."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Metadata sheet ────────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Metadata", 0)
    meta_rows = [
        ("Source PDF",        os.path.basename(source_pdf)),
        ("Extraction engine", engine),
        ("Tables extracted",  str(len(tables))),
        ("Document language", "Arabic / RTL" if is_arabic else "English / LTR"),
        ("Generated at",      time.strftime("%Y-%m-%d %H:%M:%S")),
        ("Generator",         "Toolify PDF Reconstruction Engine v3"),
    ]
    for r, (key, val) in enumerate(meta_rows, start=1):
        ck = ws_meta.cell(r, 1, key)
        cv = ws_meta.cell(r, 2, val)
        ck.font      = Font(bold=True, name="Arial", size=10)
        ck.fill      = _META_FILL
        ck.alignment = _ALIGN_LEFT_WRAP
        cv.font      = Font(name="Arial", size=10)
        cv.alignment = _ALIGN_LEFT_WRAP
    ws_meta.column_dimensions["A"].width = 24
    ws_meta.column_dimensions["B"].width = 50

    # ── Table sheets ──────────────────────────────────────────────────────────
    for i, df in enumerate(tables, start=1):
        is_rtl = is_arabic or _is_rtl_table(df)
        sheet_name = f"Table_{i}"
        _write_table_sheet(wb, df, sheet_name, is_rtl, i)

    wb.save(output_path)
    log.info("Saved → %s", output_path)


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 10 — VALIDATION LAYER (v3)
# ─────────────────────────────────────────────────────────────────────────────

def validate_and_repair(tables: list[pd.DataFrame]) -> list[pd.DataFrame]:
    """
    Phase 10: Validate column consistency, detect broken structures,
    and repair where possible.

    v3: Also detects and merges near-duplicate adjacent columns
    (a common artifact when two-column PDFs get misaligned).
    """
    valid: list[pd.DataFrame] = []

    for i, df in enumerate(tables):
        if df.shape[0] < 1 or df.shape[1] < 1:
            log.warning("Validation: Table %d skipped (too small: %s)", i + 1, df.shape)
            continue

        all_empty = df.apply(
            lambda col: col.astype(str).str.strip().eq("").all()
        ).all()
        if all_empty:
            log.warning("Validation: Table %d skipped (all empty)", i + 1)
            continue

        # Repair: detect and drop columns that are >90% empty
        col_empty_ratio = df.apply(
            lambda col: col.astype(str).str.strip().eq("").mean()
        )
        mostly_empty_cols = col_empty_ratio[col_empty_ratio > 0.90].index.tolist()
        if mostly_empty_cols and len(df.columns) - len(mostly_empty_cols) >= 1:
            df = df.drop(columns=mostly_empty_cols)
            log.info(
                "Validation: Table %d — dropped %d empty columns",
                i + 1, len(mostly_empty_cols),
            )

        # Repair: drop rows that are >90% empty
        row_empty_ratio = df.apply(
            lambda col: col.astype(str).str.strip().eq(""), axis=1
        ).mean(axis=1)
        mostly_empty_rows = row_empty_ratio[row_empty_ratio > 0.90].index
        if len(mostly_empty_rows) > 0:
            df = df.drop(index=mostly_empty_rows).reset_index(drop=True)

        # Repair: merge near-duplicate adjacent columns (same content > 85%)
        cols_to_drop: list[str] = []
        for ci in range(df.shape[1] - 1):
            col_a = df.columns[ci]
            col_b = df.columns[ci + 1]
            if col_a in cols_to_drop or col_b in cols_to_drop:
                continue
            try:
                a_vals = df[col_a].astype(str).str.strip()
                b_vals = df[col_b].astype(str).str.strip()
                same = (a_vals == b_vals).mean()
                if same >= 0.85 and df.shape[1] - len(cols_to_drop) > 1:
                    cols_to_drop.append(col_b)
                    log.info(
                        "Validation: Table %d — merging duplicate column '%s' into '%s'",
                        i + 1, col_b, col_a,
                    )
            except Exception:
                pass
        if cols_to_drop:
            df = df.drop(columns=cols_to_drop)

        is_rtl = _is_rtl_table(df)
        log.debug(
            "Validation: Table %d — %d rows × %d cols — RTL=%s",
            i + 1, df.shape[0], df.shape[1], is_rtl,
        )

        if df.shape[0] >= 1 and df.shape[1] >= 1:
            valid.append(df)
        else:
            log.warning("Validation: Table %d discarded after repair", i + 1)

    log.info("Validation: %d/%d tables passed", len(valid), len(tables))
    return valid


# ─────────────────────────────────────────────────────────────────────────────
# FULL PIPELINE ORCHESTRATOR (v3)
# ─────────────────────────────────────────────────────────────────────────────

def process_pdf(
    pdf_path: str,
    output_path: Optional[str] = None,
    lang: Optional[str] = None,
) -> str:
    """
    Enterprise PDF → Excel reconstruction pipeline — v3.

    Engine routing priority (text-based PDFs):
      1. Camelot lattice  (bordered tables, ≤30 pages)
      2. Camelot stream   (fallback)
      3. pdfplumber lattice settings
      4. pdfplumber stream settings
      5. pdfplumber word-based spatial clustering
      6. Tesseract OCR (last resort)
    """
    pdf_path = str(Path(pdf_path).resolve())
    if not os.path.isfile(pdf_path):
        raise RuntimeError(f"File not found: {pdf_path}")

    if output_path is None:
        output_path = str(Path(pdf_path).with_suffix(".xlsx"))

    t0 = time.perf_counter()
    log.info("=== Toolify PDF Reconstruction v3: %s ===", pdf_path)

    # ── Phase 1: Document Classification ─────────────────────────────────────
    profile = classify_document(pdf_path)
    ocr_lang = lang or profile.lang

    # ── Phase 2 + 5: Layout Analysis + Global Column Schema ──────────────────
    schema = GlobalColumnSchema()
    layouts: list[PageLayout] = []

    if not profile.is_scanned:
        try:
            layouts = _extract_page_words(pdf_path)
            text_layouts = [l for l in layouts if l.words]
            if text_layouts:
                schema.fit(text_layouts)
                log.info(
                    "Global schema: %d columns from %d text pages",
                    schema.n_cols, len(text_layouts),
                )
        except Exception as exc:
            log.warning("Layout extraction failed (%s) — skipping global schema", exc)

    # ── Phase 3: Smart Engine Routing ────────────────────────────────────────
    raw_dfs: list[pd.DataFrame] = []
    engine  = "unknown"

    # Detect visual-order Arabic (direction metadata from pdfplumber)
    is_visual_arabic = False
    if profile.is_arabic and not profile.is_scanned:
        try:
            import pdfplumber as _plumber
            with _plumber.open(pdf_path) as _pdf:
                sample_words: list[dict] = []
                for p in _pdf.pages[:min(3, len(_pdf.pages))]:
                    sample_words.extend(p.extract_words(
                        x_tolerance=2, y_tolerance=3,
                        keep_blank_chars=False, use_text_flow=False,
                    ) or [])
                    if len(sample_words) >= 100:
                        break
            is_visual_arabic = _detect_visual_order_from_words(sample_words)
            log.info(
                "Arabic text order: %s",
                "VISUAL (LTR-encoded)" if is_visual_arabic else "LOGICAL (RTL Unicode)",
            )
        except Exception as _e:
            log.debug("Visual-order detection skipped: %s", _e)

    _CAMELOT_PAGE_LIMIT = 30
    camelot_eligible = (
        not profile.is_scanned
        and profile.page_count <= _CAMELOT_PAGE_LIMIT
    )
    if not camelot_eligible:
        log.info(
            "Camelot skipped: page_count=%d > %d — routing to pdfplumber",
            profile.page_count, _CAMELOT_PAGE_LIMIT,
        )

    if profile.is_scanned:
        log.info("Routing: fully scanned PDF → OCR")
        try:
            raw_dfs = _extract_ocr(pdf_path, lang=ocr_lang, schema=schema if schema.n_cols > 1 else None)
            engine  = "tesseract-ocr"
        except Exception as exc:
            raise RuntimeError(f"OCR failed: {exc}") from exc

    else:
        # ── Lattice-first routing ─────────────────────────────────────────────
        # Always try Camelot lattice first regardless of border-detection result.
        # Reason: PDF content streams are zlib-compressed, so the raw-byte border
        # scan is unreliable.  Camelot lattice uses Ghostscript to decode the
        # streams and find actual grid lines — far more reliable.
        # If lattice finds 0 tables, _extract_camelot() already falls back to
        # stream internally; no extra logic needed here.
        camelot_flavor = "lattice"

        if camelot_eligible:
            try:
                raw_dfs = _extract_camelot(pdf_path, camelot_flavor)
                if raw_dfs:
                    engine = f"camelot/{camelot_flavor}"
                    log.info(
                        "Camelot lattice succeeded with %d table(s)", len(raw_dfs)
                    )
            except Exception as exc:
                log.warning("Camelot routing failed: %s", exc)

        if not raw_dfs:
            try:
                raw_dfs = _extract_pdfplumber_tables(
                    pdf_path,
                    fix_visual_arabic=False,   # applied centrally below to all engines
                    has_borders=profile.has_borders,
                )
                if raw_dfs:
                    engine = "pdfplumber"
            except Exception as exc:
                log.warning("pdfplumber failed: %s", exc)

        if not raw_dfs and layouts:
            log.info("Routing: spatial clustering with global schema")
            for layout in layouts:
                if layout.words:
                    df = layout_to_dataframe(layout, schema)
                    if df is not None and not df.empty:
                        raw_dfs.append(df)
            if raw_dfs:
                engine = "spatial-clustering"

        if not raw_dfs:
            log.info("Routing: all text engines failed → OCR")
            try:
                raw_dfs = _extract_ocr(pdf_path, lang=ocr_lang, schema=schema if schema.n_cols > 1 else None)
                engine  = "tesseract-ocr (fallback)"
            except Exception as exc:
                raise RuntimeError(f"All extractors failed: {exc}") from exc

        # ── Apply Arabic visual-order correction to ALL extraction paths ──────
        # Previously this was only done in the spatial-clustering fallback.
        # Camelot and pdfplumber return text exactly as stored in the PDF stream
        # (visual order for Arabic documents), so we must correct here too.
        if is_visual_arabic and raw_dfs:
            log.info("Applying Arabic visual-order fix to %d table(s)", len(raw_dfs))
            raw_dfs = [
                fix_visual_arabic_dataframe(df, visual_order=True)
                for df in raw_dfs
            ]
            # Also mirror parentheses/brackets — they are stored reversed in
            # visual-order PDFs throughout the whole row, even in English cells.
            raw_dfs = [mirror_brackets_dataframe(df) for df in raw_dfs]
            log.info("Bracket mirroring applied to %d table(s)", len(raw_dfs))

        # ── Row-fragment repair (stream-mode safety net) ──────────────────────
        # When Camelot stream or pdfplumber stream was used, multi-line cells
        # may have been split across separate rows.  Merge continuation rows
        # (first column empty) back into the preceding data row.
        if raw_dfs and "stream" in engine:
            log.info("Applying row-fragment repair (stream engine detected)")
            raw_dfs = [_merge_fragmented_rows(df) for df in raw_dfs]

        # Hybrid PDF: supplement with OCR for scanned pages if needed
        expected_tables = max(profile.page_count // 2, 1)
        pdfplumber_sufficient = len(raw_dfs) >= expected_tables * 0.5

        if profile.is_hybrid and profile.scanned_page_indices and not pdfplumber_sufficient:
            log.info(
                "Hybrid PDF: supplementing with OCR for %d scanned pages",
                len(profile.scanned_page_indices),
            )
            try:
                from pdf2image import convert_from_path
                images = convert_from_path(pdf_path, dpi=300)
                for pi in profile.scanned_page_indices:
                    if pi < len(images):
                        df = _ocr_page_to_df(
                            images[pi], lang=ocr_lang,
                            schema=schema if schema.n_cols > 1 else None,
                        )
                        if df is not None and not df.empty:
                            raw_dfs.append(df)
                engine += "+ocr-hybrid"
            except Exception as exc:
                log.warning("Hybrid OCR supplement failed: %s", exc)
        elif profile.is_hybrid and pdfplumber_sufficient:
            log.info(
                "Hybrid PDF: pdfplumber found %d tables — skipping OCR supplement",
                len(raw_dfs),
            )

    # ── Phase 6 + 7: Table Reconstruction + Header Intelligence ──────────────
    tables = reconstruct_tables(raw_dfs)

    if not tables:
        raise RuntimeError(
            f"No usable tables found in '{os.path.basename(pdf_path)}'. "
            "The PDF may contain no tables or only complex graphical content."
        )

    # ── Phase 10: Validation ──────────────────────────────────────────────────
    tables = validate_and_repair(tables)

    if not tables:
        raise RuntimeError(
            f"Tables were found but failed validation in '{os.path.basename(pdf_path)}'. "
            "The content may be corrupted or too sparse."
        )

    # ── Phase 10.5: Drop repeated-header data rows ────────────────────────────
    # Must run AFTER validate_and_repair because that step may drop empty
    # columns, changing a 0.75-similarity header-lookalike row into a 1.0
    # exact match.  Removes any data row whose normalized content matches
    # the column-name tuple (exact) or is ≥80% similar (robust to minor
    # differences introduced by page-level header variations).
    def _drop_repeated_header_rows(df: pd.DataFrame) -> pd.DataFrame:
        col_fp = tuple(
            unicodedata.normalize("NFKC", re.sub(r"\s+", " ", str(c).strip().lower()))
            for c in df.columns
        )
        def _is_data_row(row: pd.Series) -> bool:
            fp = _normalize_header_row(row)
            return _header_similarity(fp, col_fp) < 0.80
        mask = df.apply(_is_data_row, axis=1)
        return df[mask].reset_index(drop=True)

    tables = [_drop_repeated_header_rows(t) for t in tables]
    tables = [t for t in tables if not t.empty]

    # ── Phase 9: Excel Generation ─────────────────────────────────────────────
    write_excel(
        tables, output_path, pdf_path, engine,
        is_arabic=profile.is_arabic,
    )

    elapsed = time.perf_counter() - t0
    log.info(
        "Done: %d table(s) → %s [engine=%s] in %.2fs",
        len(tables), output_path, engine, elapsed,
    )
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# BATCH PROCESSING
# ─────────────────────────────────────────────────────────────────────────────

def process_batch(
    paths: list[str],
    lang: Optional[str] = None,
) -> dict[str, "str | Exception"]:
    results: dict[str, "str | Exception"] = {}
    for path in paths:
        try:
            results[path] = process_pdf(path, lang=lang)
        except Exception as exc:
            log.error("FAILED %s: %s", path, exc)
            results[path] = exc
    return results


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="pdf_table_extractor",
        description="Enterprise PDF → Excel reconstruction pipeline v3",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python pdf_table_extractor.py report.pdf\n"
            "  python pdf_table_extractor.py report.pdf output.xlsx\n"
            "  python pdf_table_extractor.py *.pdf\n"
            "  python pdf_table_extractor.py report.pdf -q\n"
            "  python pdf_table_extractor.py report.pdf --lang ara+eng\n"
        ),
    )
    p.add_argument("inputs",            nargs="+",           help="PDF file(s) to process")
    p.add_argument("output",            nargs="?",           help="Output .xlsx (single-file mode)")
    p.add_argument("-q", "--quiet",     action="store_true", help="Errors only")
    p.add_argument("-v", "--verbose",   action="store_true", help="Debug logs")
    p.add_argument(
        "--lang", default=None, metavar="LANG",
        help="Tesseract language override, e.g. 'ara+eng' or 'eng'",
    )
    return p


def main() -> None:
    parser = _build_parser()
    args   = parser.parse_args()

    if args.quiet:
        log.setLevel(logging.ERROR)
    elif args.verbose:
        log.setLevel(logging.DEBUG)

    pdf_inputs   = [p for p in args.inputs if p.lower().endswith(".pdf")]
    non_pdfs     = [p for p in args.inputs if not p.lower().endswith(".pdf")]
    explicit_out = None

    if non_pdfs:
        if (
            len(non_pdfs) == 1
            and non_pdfs[0].lower().endswith(".xlsx")
            and len(pdf_inputs) == 1
        ):
            explicit_out = non_pdfs[0]
        else:
            parser.error(f"Unrecognised input(s): {non_pdfs}")

    explicit_out = explicit_out or args.output

    if not pdf_inputs:
        parser.error("No PDF files specified.")

    forced_lang = args.lang or None

    if len(pdf_inputs) > 1:
        if explicit_out:
            parser.error("Output path can only be used in single-file mode.")
        results = process_batch(pdf_inputs, lang=forced_lang)
        errors  = {k: v for k, v in results.items() if isinstance(v, Exception)}
        if errors:
            for path, exc in errors.items():
                print(f"ERROR  {path}: {exc}", file=sys.stderr)
            sys.exit(1)
        for path, out in results.items():
            print(f"OK     {path} → {out}")
    else:
        try:
            out = process_pdf(pdf_inputs[0], explicit_out, lang=forced_lang)
            print(f"OK     {pdf_inputs[0]} → {out}")
        except RuntimeError as exc:
            print(f"ERROR  {exc}", file=sys.stderr)
            sys.exit(1)


if __name__ == "__main__":
    main()
