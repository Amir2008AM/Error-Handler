#!/usr/bin/env python3
"""
PDF → Excel Multi-Engine Table Extractor v2
============================================
Engine cascade (highest accuracy first):
  1. Camelot lattice  — vector-line bordered tables
  2. Camelot stream   — whitespace-inferred column layout
  3. pdfplumber       — character-level table detection
  4. PyMuPDF blocks   — position-based text layout

Output: .xlsx with:
  • One sheet per page (or single Sheet1 for one-page PDFs)
  • Auto-sized column widths (capped at 60 chars)
  • Bold dark-header row (auto-detected)
  • RTL worksheet direction when Arabic text found
  • Frozen first row (header)
  • Proper UTF-8 encoding — no character corruption
"""
from __future__ import annotations

import sys
import os
import re
import argparse
import traceback
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Arabic / RTL detection ────────────────────────────────────────────────────
_AR_RE = re.compile(
    r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF'
    r'\uFB50-\uFDFF\uFE70-\uFEFF\u0660-\u0669\u06F0-\u06F9]'
)


def _has_arabic(text: str) -> bool:
    return bool(_AR_RE.search(str(text or '')))


def _table_has_arabic(data: list[list]) -> bool:
    for row in data:
        for cell in row:
            if _has_arabic(str(cell or '')):
                return True
    return False


# ── Cell cleaning ─────────────────────────────────────────────────────────────
def _clean(val) -> str:
    if val is None:
        return ''
    s = str(val)
    # Strip control characters except LF/CR/TAB
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s)
    return s.strip()


def _normalize_row_lengths(data: list[list]) -> list[list[str]]:
    """Ensure all rows have the same length; clean every cell."""
    if not data:
        return []
    max_cols = max((len(r) for r in data), default=0)
    result = []
    for row in data:
        cleaned = [_clean(c) for c in row]
        cleaned += [''] * (max_cols - len(cleaned))
        result.append(cleaned)
    return result


# ── Excel formatting helpers ───────────────────────────────────────────────────
_HEADER_FILL   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
_SUBHDR_FILL   = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
_HEADER_FONT   = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
_BODY_FONT     = Font(size=10, name='Calibri')
_THIN_SIDE     = Side(style='thin',   color='BFBFBF')
_THICK_SIDE    = Side(style='medium', color='1F4E79')
_CELL_BORDER   = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                         top=_THIN_SIDE,  bottom=_THIN_SIDE)
_HEADER_BORDER = Border(left=_THICK_SIDE, right=_THICK_SIDE,
                         top=_THICK_SIDE,  bottom=_THICK_SIDE)

# Alternating row fills
_ALT_FILL = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')


def _is_header_candidate(row: list[str]) -> bool:
    """Heuristic: all cells non-empty and no purely numeric values."""
    non_empty = [c for c in row if c.strip()]
    if not non_empty or len(non_empty) < len(row) // 2:
        return False
    numeric = sum(1 for c in non_empty if re.match(r'^[\d\s\.,٠-٩\-\+\(\)%]+$', c))
    return numeric < len(non_empty) // 2


def write_sheet(ws, data: list[list[str]], is_rtl: bool = False) -> None:
    """Write a 2D list of clean strings to the worksheet with full formatting."""
    if not data:
        return

    if is_rtl:
        ws.sheet_view.rightToLeft = True

    detect_header = _is_header_candidate(data[0]) if data else False
    col_widths: dict[int, float] = {}

    for r_idx, row in enumerate(data, start=1):
        is_header_row = (r_idx == 1 and detect_header)
        is_alt        = (not is_header_row) and (r_idx % 2 == 0)

        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val or None)

            # ── Font & fill ───────────────────────────────────────────────
            if is_header_row:
                cell.font   = _HEADER_FONT
                cell.fill   = _HEADER_FILL
                cell.border = _HEADER_BORDER
            else:
                cell.font   = _BODY_FONT
                cell.border = _CELL_BORDER
                if is_alt:
                    cell.fill = _ALT_FILL

            # ── Alignment ─────────────────────────────────────────────────
            cell_ar   = _has_arabic(val)
            h_align   = 'right' if (is_rtl or cell_ar) else 'left'
            reading   = 2 if (is_rtl or cell_ar) else 1

            if is_header_row:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center',
                    wrap_text=False, readingOrder=reading,
                )
            else:
                cell.alignment = Alignment(
                    horizontal=h_align, vertical='center',
                    wrap_text=True, readingOrder=reading,
                )

            # ── Column width tracking ─────────────────────────────────────
            display_len = min(len(val), 60)
            col_widths[c_idx] = max(col_widths.get(c_idx, 6), display_len + 2)

    # Apply column widths
    for c_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = float(width)

    # Set uniform row height
    for r_idx in range(1, len(data) + 1):
        ws.row_dimensions[r_idx].height = 18

    # Freeze header row
    if detect_header and len(data) > 1:
        ws.freeze_panes = 'A2'

    # Auto-filter on header row
    if detect_header and len(data[0]) > 0:
        last_col = get_column_letter(len(data[0]))
        ws.auto_filter.ref = f'A1:{last_col}1'


# ── Engine 1: Camelot ─────────────────────────────────────────────────────────
def _camelot_extract(pdf_path: str, flavor: str) -> list[dict]:
    import camelot  # noqa: import inside function to keep startup fast

    kwargs: dict = {'suppress_stdout': True}
    if flavor == 'lattice':
        # copy_text copies the spanning-cell text into all sub-cells so no data is lost
        try:
            tables = camelot.read_pdf(
                pdf_path, pages='all', flavor='lattice',
                copy_text=['v'], **kwargs
            )
        except TypeError:
            # Older Camelot API without copy_text
            tables = camelot.read_pdf(pdf_path, pages='all', flavor='lattice', **kwargs)
    else:
        tables = camelot.read_pdf(
            pdf_path, pages='all', flavor='stream',
            edge_tol=100, row_tol=10, **kwargs
        )

    results = []
    for t in tables:
        if t.df is None or t.df.empty:
            continue
        rows = _normalize_row_lengths([list(r) for r in t.df.values.tolist()])
        non_empty = sum(1 for row in rows for c in row if c.strip())
        if non_empty < 3:   # skip near-empty tables
            continue
        results.append({
            'page':     t.page,
            'data':     rows,
            'accuracy': t.accuracy,
            'engine':   f'camelot-{flavor}',
        })

    return results


# ── Engine 2: pdfplumber ──────────────────────────────────────────────────────
def _pdfplumber_extract(pdf_path: str) -> list[dict]:
    import pdfplumber

    results = []
    strats = [
        {'vertical_strategy': 'lines_strict', 'horizontal_strategy': 'lines_strict',
         'snap_tolerance': 3, 'join_tolerance': 3},
        {'vertical_strategy': 'lines', 'horizontal_strategy': 'lines',
         'snap_tolerance': 5, 'join_tolerance': 5},
        {'vertical_strategy': 'text', 'horizontal_strategy': 'text',
         'snap_tolerance': 5, 'join_tolerance': 5, 'min_words_vertical': 3},
    ]

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            extracted = None
            for strat in strats:
                try:
                    tables = page.extract_tables(strat)
                    if tables:
                        extracted = tables
                        break
                    # Fallback: single table
                    t = page.extract_table(strat)
                    if t:
                        extracted = [t]
                        break
                except Exception:
                    continue

            if not extracted:
                continue

            for t in extracted:
                if not t:
                    continue
                rows = _normalize_row_lengths(t)
                non_empty = sum(1 for row in rows for c in row if c.strip())
                if non_empty < 3:
                    continue
                results.append({
                    'page':     page_num,
                    'data':     rows,
                    'accuracy': 70,
                    'engine':   'pdfplumber',
                })

    return results


# ── Engine 3: PyMuPDF layout ──────────────────────────────────────────────────
def _pymupdf_extract(pdf_path: str) -> list[dict]:
    """
    Position-based extraction: groups words by Y-band into rows,
    then groups rows into columns using X-clustering.
    Works for any PDF with selectable text, even without table borders.
    """
    import fitz

    Y_TOLERANCE = 4   # pixels — words within this Y-band are the same row
    X_GAP_RATIO = 0.8  # gap > 80% of average word width → new column

    doc   = fitz.open(pdf_path)
    results = []

    for page_num in range(len(doc)):
        page  = doc[page_num]
        words = page.get_text('words')  # (x0,y0,x1,y1, text, block,line,word)
        if not words:
            continue

        # Group by rounded Y
        rows_by_y: dict[int, list[tuple]] = {}
        for w in words:
            y_key = round(w[1] / Y_TOLERANCE) * Y_TOLERANCE
            rows_by_y.setdefault(y_key, []).append(w)

        if not rows_by_y:
            continue

        # Sort words within each row by X, then sort rows by Y
        row_list: list[list[str]] = []
        for y_key in sorted(rows_by_y.keys()):
            row_words = sorted(rows_by_y[y_key], key=lambda w: w[0])
            row_list.append([_clean(w[4]) for w in row_words])

        if not row_list:
            continue

        rows_norm = _normalize_row_lengths(row_list)

        non_empty = sum(1 for row in rows_norm for c in row if c.strip())
        if non_empty < 3:
            continue

        results.append({
            'page':     page_num + 1,
            'data':     rows_norm,
            'accuracy': 50,
            'engine':   'pymupdf',
        })

    doc.close()
    return results


# ── Orchestrator ──────────────────────────────────────────────────────────────
def extract_tables(pdf_path: str) -> list[dict]:
    errors: list[str] = []

    # 1. Camelot lattice — best for bordered tables
    try:
        r = _camelot_extract(pdf_path, 'lattice')
        if r:
            _log(f'Camelot lattice: {len(r)} table(s)')
            return r
    except Exception as e:
        errors.append(f'camelot-lattice: {e}')

    # 2. Camelot stream — best for whitespace tables
    try:
        r = _camelot_extract(pdf_path, 'stream')
        if r:
            _log(f'Camelot stream: {len(r)} table(s)')
            return r
    except Exception as e:
        errors.append(f'camelot-stream: {e}')

    # 3. pdfplumber
    try:
        r = _pdfplumber_extract(pdf_path)
        if r:
            _log(f'pdfplumber: {len(r)} table(s)')
            return r
    except Exception as e:
        errors.append(f'pdfplumber: {e}')

    # 4. PyMuPDF layout
    try:
        r = _pymupdf_extract(pdf_path)
        if r:
            _log(f'pymupdf: {len(r)} page(s)')
            return r
    except Exception as e:
        errors.append(f'pymupdf: {e}')

    detail = ' | '.join(errors) if errors else 'no extractable content'
    raise RuntimeError(f'All extractors failed: {detail}')


# ── Build XLSX workbook ───────────────────────────────────────────────────────
def build_xlsx(tables: list[dict], output_path: str) -> None:
    wb      = openpyxl.Workbook()
    pages   = {}
    for t in tables:
        pages.setdefault(t['page'], []).append(t)

    used_names: set[str] = set()
    first_sheet = True

    for page_num in sorted(pages.keys()):
        for t_idx, table in enumerate(pages[page_num]):
            data = table['data']
            if not data:
                continue

            # ── Sheet name ────────────────────────────────────────────────
            total_tables  = sum(len(v) for v in pages.values())
            total_pages   = len(pages)
            if total_tables == 1:
                raw_name = 'Sheet1'
            elif total_pages == 1:
                raw_name = f'Table {t_idx + 1}'
            elif len(pages[page_num]) == 1:
                raw_name = f'Page {page_num}'
            else:
                raw_name = f'Page {page_num} T{t_idx + 1}'

            base = raw_name[:28]
            name = base
            suf  = 2
            while name in used_names:
                name = f'{base[:25]}_{suf}'; suf += 1
            used_names.add(name)

            # ── Create sheet ──────────────────────────────────────────────
            if first_sheet:
                ws = wb.active
                ws.title = name
                first_sheet = False
            else:
                ws = wb.create_sheet(name)

            is_rtl = _table_has_arabic(data)
            write_sheet(ws, data, is_rtl=is_rtl)

    if not used_names:
        raise RuntimeError('No tables were written to the workbook')

    wb.save(output_path)
    _log(f'Saved → {output_path}')


# ── Logging ───────────────────────────────────────────────────────────────────
_quiet = False

def _log(msg: str) -> None:
    if not _quiet:
        print(f'[pdf-to-excel] {msg}', file=sys.stderr, flush=True)


# ── CLI ───────────────────────────────────────────────────────────────────────
def main() -> None:
    global _quiet
    ap = argparse.ArgumentParser(description='PDF → Excel table extractor')
    ap.add_argument('pdf',  help='Input PDF file')
    ap.add_argument('xlsx', help='Output .xlsx file')
    ap.add_argument('-q', '--quiet', action='store_true', help='Suppress log output')
    args = ap.parse_args()

    if args.quiet:
        _quiet = True

    pdf_path  = Path(args.pdf)
    xlsx_path = Path(args.xlsx)

    if not pdf_path.exists():
        print(f'File not found: {pdf_path}', file=sys.stderr)
        sys.exit(1)

    try:
        tables = extract_tables(str(pdf_path))
        build_xlsx(tables, str(xlsx_path))
    except Exception as exc:
        # Print clean error to stderr so the Node caller can surface it
        print(str(exc), file=sys.stderr, flush=True)
        if not _quiet:
            traceback.print_exc(file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
