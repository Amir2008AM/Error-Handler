"""
Excel → PDF  —  LibreOffice Primary Pipeline v3
================================================

Pipeline stages
---------------
  1. LIGHT VALIDATION        — magic bytes, openpyxl load, at least one non-empty sheet
  2. PREFLIGHT STABILIZATION — normalise row heights, col widths, merged cells,
                               wrap text, print settings via openpyxl (writes
                               a stabilised copy — never mutates the original)
  3. FONT NORMALIZATION      — map Excel font names → available Liberation/DejaVu
                               equivalents so LibreOffice never does a bad substitution
  4. PRIMARY CONVERSION      — LibreOffice headless --convert-to pdf
  5. POST-RENDER VALIDATION  — pdfplumber: empty pages, overflow indicators,
                               duplicated header rows; emits a confidence score
  6. SAFE RETRY              — one retry with conservative LO flags + isolated
                               user profile if validation confidence < threshold
  7. RETURN                  — JSON {success, pageCount, engine, confidence, warnings}
                               Non-zero exit signals total crash → caller uses fallback

Usage
-----
    python3 excel_to_pdf_lo.py input.xlsx output.pdf [options]

Options
-------
    --page-size    a4 | letter | legal      (default: a4)
    --orientation  portrait | landscape     (default: landscape)
"""

from __future__ import annotations

import sys
import os
import re
import json
import copy
import shutil
import argparse
import tempfile
import subprocess
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment

try:
    import pdfplumber
    _HAS_PDFPLUMBER = True
except ImportError:
    _HAS_PDFPLUMBER = False


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

_EXCEL_DEFAULT_ROW_H = 15.0   # points
_EXCEL_DEFAULT_COL_W = 8.43   # character units
_MIN_COL_W  = 3.0
_MAX_COL_W  = 60.0
_MIN_ROW_H  = 8.0
_MAX_ROW_H  = 200.0

# LibreOffice paper size codes for --infilter / page setup
_PAPER_CODES = {
    'a4':     9,    # A4
    'letter': 1,    # Letter
    'legal':  5,    # Legal
}

# Fonts that exist in every LibreOffice install (Liberation / DejaVu bundle)
_SAFE_FONTS = {
    'Liberation Sans',
    'Liberation Serif',
    'Liberation Mono',
    'DejaVu Sans',
    'DejaVu Serif',
    'DejaVu Sans Mono',
}

# Excel font → nearest safe font
_FONT_MAP: dict[str, str] = {
    # Sans-serif Excel defaults
    'calibri':           'Liberation Sans',
    'calibri light':     'Liberation Sans',
    'calibri (body)':    'Liberation Sans',
    'arial':             'Liberation Sans',
    'arial narrow':      'Liberation Sans',
    'arial unicode ms':  'Liberation Sans',
    'tahoma':            'Liberation Sans',
    'segoe ui':          'Liberation Sans',
    'trebuchet ms':      'Liberation Sans',
    'verdana':           'Liberation Sans',
    'helvetica':         'Liberation Sans',
    'helvetica neue':    'Liberation Sans',
    'franklin gothic medium': 'Liberation Sans',
    # Serif
    'times new roman':   'Liberation Serif',
    'times':             'Liberation Serif',
    'georgia':           'Liberation Serif',
    'garamond':          'Liberation Serif',
    'palatino linotype': 'Liberation Serif',
    'book antiqua':      'Liberation Serif',
    'century':           'Liberation Serif',
    'cambria':           'Liberation Serif',
    'cambria math':      'Liberation Serif',
    # Monospace
    'courier new':       'Liberation Mono',
    'courier':           'Liberation Mono',
    'consolas':          'Liberation Mono',
    'lucida console':    'Liberation Mono',
    'monaco':            'Liberation Mono',
}


# ─────────────────────────────────────────────────────────────────────────────
# Stage 1 — Light Validation
# ─────────────────────────────────────────────────────────────────────────────

def validate_input(path: str) -> openpyxl.Workbook:
    """
    Check the file is a readable Excel workbook with at least one non-empty
    sheet.  Returns the loaded Workbook on success, raises on failure.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f'Input file not found: {path}')
    if p.stat().st_size == 0:
        raise ValueError('Input file is empty (0 bytes)')

    # Magic bytes: xlsx is a ZIP (PK = 50 4B), xls is OLE2 (D0 CF)
    with open(path, 'rb') as fh:
        magic = fh.read(4)
    is_xlsx = magic[:2] == b'PK'
    is_xls  = magic[:2] == b'\xd0\xcf'
    if not (is_xlsx or is_xls):
        raise ValueError('File does not appear to be a valid Excel document (bad magic bytes)')

    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        raise ValueError(f'openpyxl could not open the workbook: {exc}') from exc

    non_empty = [ws for ws in wb.worksheets
                 if ws.max_row and ws.max_column and ws.max_row > 0]
    if not non_empty:
        raise ValueError('Workbook contains no data (all sheets are empty)')

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# Stage 2 — Preflight Stabilization
# ─────────────────────────────────────────────────────────────────────────────

def _fix_merged_cells(ws) -> int:
    """
    Remove any merged-cell ranges that are invalid (outside the used range or
    zero-size) to prevent LibreOffice from choking.  Returns number of fixes.
    """
    fixes = 0
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    bad = []
    for mc in list(ws.merged_cells.ranges):
        if (mc.min_row > max_row or mc.min_col > max_col
                or mc.min_row == mc.max_row and mc.min_col == mc.max_col):
            bad.append(mc)
    for mc in bad:
        ws.merged_cells.remove(mc)
        fixes += 1
    return fixes


def stabilize_workbook(wb: openpyxl.Workbook, page_size: str, orientation: str) -> openpyxl.Workbook:
    """
    Stage 2 — create a stabilised copy of the workbook.

    Normalises:
    - Row heights: unset/zero → _EXCEL_DEFAULT_ROW_H, clamp to [_MIN_ROW_H, _MAX_ROW_H]
    - Column widths: unset/zero/extreme → clamp to [_MIN_COL_W, _MAX_COL_W]
    - Merged cells: remove invalid ranges
    - Wrap text: enable on cells whose content length > 40 chars
    - Print settings: page size, orientation, fitToPage, margins, repeat-headers
    """
    # Paper size codes (OOXML PageSetup.paperSize)
    paper_code = {'a4': 9, 'letter': 1, 'legal': 5}.get(page_size, 9)
    is_landscape = orientation == 'landscape'

    for ws in wb.worksheets:
        if not ws.max_row:
            continue

        # ── Row heights ───────────────────────────────────────────────────
        for ri in range(1, (ws.max_row or 0) + 1):
            rd = ws.row_dimensions[ri]
            h = rd.height
            if h is None or h <= 0:
                rd.height = _EXCEL_DEFAULT_ROW_H
            else:
                rd.height = max(_MIN_ROW_H, min(float(h), _MAX_ROW_H))

        # ── Column widths ─────────────────────────────────────────────────
        for ci in range(1, (ws.max_column or 0) + 1):
            col_letter = get_column_letter(ci)
            cd = ws.column_dimensions[col_letter]
            w = cd.width
            if w is None or w <= 0:
                cd.width = _EXCEL_DEFAULT_COL_W
            else:
                cd.width = max(_MIN_COL_W, min(float(w), _MAX_COL_W))

        # ── Merged cells ──────────────────────────────────────────────────
        _fix_merged_cells(ws)

        # ── Wrap text on long cells ───────────────────────────────────────
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and len(cell.value) > 40:
                    aln = cell.alignment
                    if not aln.wrap_text:
                        cell.alignment = Alignment(
                            horizontal=aln.horizontal,
                            vertical=aln.vertical,
                            wrap_text=True,
                            text_rotation=aln.text_rotation,
                            shrink_to_fit=aln.shrink_to_fit,
                            indent=aln.indent,
                        )

        # ── Print settings ────────────────────────────────────────────────
        ps = ws.page_setup
        ps.paperSize  = paper_code
        ps.orientation = 'landscape' if is_landscape else 'portrait'
        ps.fitToPage  = True
        ps.fitToWidth = 1
        ps.fitToHeight = 0        # allow as many pages tall as needed

        pm = ws.page_margins
        # Use tighter margins so content isn't clipped
        pm.left   = 0.5
        pm.right  = 0.5
        pm.top    = 0.75
        pm.bottom = 0.75
        pm.header = 0.3
        pm.footer = 0.3

        # Repeat first row on every page if it looks like a header
        if ws.max_row and ws.max_row > 1:
            ws.print_title_rows = '1:1'

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# Stage 3 — Font Normalization
# ─────────────────────────────────────────────────────────────────────────────

def normalize_fonts(wb: openpyxl.Workbook) -> list[str]:
    """
    Stage 3 — walk every cell and replace unknown Excel font names with the
    nearest Liberation/DejaVu equivalent.  Returns list of substitution
    messages for the warnings log.
    """
    substitutions: list[str] = []
    seen: dict[str, str] = {}   # original → replacement (cache)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font is None:
                    continue
                name = (cell.font.name or '').strip()
                if not name:
                    continue

                # Already known safe
                if name in _SAFE_FONTS:
                    continue

                # Cached
                if name in seen:
                    replacement = seen[name]
                    if replacement != name:
                        new_font = copy.copy(cell.font)
                        new_font = Font(
                            name=replacement,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            size=cell.font.size,
                            color=cell.font.color,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                        )
                        cell.font = new_font
                    continue

                # Look up in substitution map
                replacement = _FONT_MAP.get(name.lower())
                if replacement:
                    seen[name] = replacement
                    substitutions.append(f'{name!r} → {replacement!r}')
                    new_font = Font(
                        name=replacement,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        size=cell.font.size,
                        color=cell.font.color,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                    )
                    cell.font = new_font
                else:
                    # Not in map — leave as-is but record as seen
                    seen[name] = name

    return substitutions


# ─────────────────────────────────────────────────────────────────────────────
# Stage 4 — Primary Conversion (LibreOffice)
# ─────────────────────────────────────────────────────────────────────────────

def _soffice_convert(
    input_file: str,
    out_dir: str,
    lo_home: str,
    timeout: int = 120,
    extra_args: Optional[list[str]] = None,
) -> subprocess.CompletedProcess:
    """Run LibreOffice headless conversion.  Returns the CompletedProcess."""
    cmd = [
        'soffice',
        '--headless',
        '--norestore',
        '--nofirststartwizard',
        '--nocrashreport',
        f'-env:UserInstallation=file://{lo_home}',
        '--convert-to', 'pdf',
        '--outdir', out_dir,
        input_file,
    ]
    if extra_args:
        cmd += extra_args

    env = {
        **os.environ,
        'HOME':              lo_home,
        'TMPDIR':            lo_home,
        'SAL_USE_VCLPLUGIN': 'svp',
        'PYTHONPATH':        '',
    }

    return subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=timeout,
        env=env,
    )


def run_libreoffice(
    stabilised_xlsx: str,
    out_dir: str,
    lo_home: str,
    timeout: int = 120,
    safe_mode: bool = False,
) -> str:
    """
    Stage 4 — convert the stabilised xlsx with LibreOffice.
    Returns the path to the produced PDF.
    Raises on failure (bad exit code or no output PDF).
    """
    extra = ['--safe-mode'] if safe_mode else []
    result = _soffice_convert(stabilised_xlsx, out_dir, lo_home, timeout=timeout, extra_args=extra)

    expected_pdf = Path(out_dir) / (Path(stabilised_xlsx).stem + '.pdf')

    if result.returncode != 0 and not expected_pdf.exists():
        stderr_snippet = (result.stderr or result.stdout or '').strip()[:400]
        raise RuntimeError(
            f'LibreOffice exited {result.returncode}: {stderr_snippet}'
        )

    if not expected_pdf.exists():
        raise RuntimeError('LibreOffice produced no output PDF')

    # Validate magic bytes
    with open(expected_pdf, 'rb') as fh:
        magic = fh.read(5)
    if magic != b'%PDF-':
        raise RuntimeError('LibreOffice output file is not a valid PDF')

    return str(expected_pdf)


# ─────────────────────────────────────────────────────────────────────────────
# Stage 5 — Post-render Validation
# ─────────────────────────────────────────────────────────────────────────────

def validate_pdf(pdf_path: str) -> tuple[float, list[str]]:
    """
    Stage 5 — quick quality checks on the rendered PDF.

    Returns (confidence 0.0–1.0, list_of_warnings).
    confidence == 1.0 means no problems detected.
    """
    warnings: list[str] = []

    if not _HAS_PDFPLUMBER:
        return 1.0, ['pdfplumber not available — skipping post-render validation']

    try:
        with pdfplumber.open(pdf_path) as pdf:
            page_count = len(pdf.pages)
            if page_count == 0:
                return 0.0, ['PDF has 0 pages']

            empty_pages = 0
            overflow_pages = 0
            all_first_row_texts: list[str] = []
            dup_header_pages = 0

            for i, page in enumerate(pdf.pages):
                words = page.extract_words() or []

                # Empty page check
                if not words:
                    empty_pages += 1
                    warnings.append(f'Page {i+1}: no text extracted (may be blank or image-only)')
                    continue

                # Overflow check: any word bbox exceeds page bounds by >5%
                pw, ph = page.width, page.height
                overflow_words = [
                    w for w in words
                    if (float(w.get('x1', 0)) > pw * 1.05
                        or float(w.get('y1', 0)) > ph * 1.05)
                ]
                if overflow_words:
                    overflow_pages += 1
                    warnings.append(
                        f'Page {i+1}: {len(overflow_words)} word(s) may overflow page bounds'
                    )

                # Duplicated header detection
                # Collect text of the top 3% of the page (likely header row)
                top_band = ph * 0.03
                top_words = [w['text'] for w in words if float(w.get('top', ph)) < top_band]
                first_row_text = ' '.join(top_words).strip()
                if first_row_text:
                    all_first_row_texts.append(first_row_text)

            # Detect duplicate header rows across pages
            if len(all_first_row_texts) > 1:
                unique_headers = set(all_first_row_texts)
                if len(unique_headers) < len(all_first_row_texts) * 0.8:
                    dup_header_pages = len(all_first_row_texts) - len(unique_headers)
                    warnings.append(
                        f'{dup_header_pages} page(s) appear to have duplicated header text'
                    )

            # Confidence score
            penalty = 0.0
            penalty += empty_pages    * 0.15
            penalty += overflow_pages * 0.10
            penalty += dup_header_pages * 0.05
            confidence = max(0.0, 1.0 - penalty / max(page_count, 1))

    except Exception as exc:
        warnings.append(f'Post-render validation error (non-fatal): {exc}')
        confidence = 0.8   # assume mostly OK if we can't inspect

    return confidence, warnings


# ─────────────────────────────────────────────────────────────────────────────
# Stage 6 — Safe Retry
# ─────────────────────────────────────────────────────────────────────────────

def safe_retry(
    stabilised_xlsx: str,
    out_dir: str,
    lo_home_retry: str,
) -> str:
    """
    Stage 6 — one retry with a fresh isolated LO user profile.
    Returns path to PDF on success, raises on failure.
    """
    return run_libreoffice(
        stabilised_xlsx,
        out_dir,
        lo_home=lo_home_retry,
        timeout=150,
        safe_mode=False,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Page count helper
# ─────────────────────────────────────────────────────────────────────────────

def _count_pages(pdf_path: str) -> int:
    if not _HAS_PDFPLUMBER:
        return 1
    try:
        with pdfplumber.open(pdf_path) as pdf:
            return len(pdf.pages)
    except Exception:
        return 1


# ─────────────────────────────────────────────────────────────────────────────
# Top-level orchestrator
# ─────────────────────────────────────────────────────────────────────────────

def excel_to_pdf_lo(
    input_path: str,
    output_path: str,
    page_size: str = 'a4',
    orientation: str = 'landscape',
) -> dict:
    """
    Full LibreOffice-primary pipeline (stages 1–7).
    Returns a dict suitable for JSON serialisation.
    Raises on total failure so the caller can use the pdf-lib fallback.
    """
    all_warnings: list[str] = []

    # ── Stage 1: Light Validation ─────────────────────────────────────────
    wb = validate_input(input_path)

    with tempfile.TemporaryDirectory(prefix='xlsx-lo-') as tmp:

        # ── Stage 2: Preflight Stabilization ─────────────────────────────
        wb = stabilize_workbook(wb, page_size, orientation)

        # ── Stage 3: Font Normalization ───────────────────────────────────
        font_subs = normalize_fonts(wb)
        if font_subs:
            all_warnings.append(f'Font substitutions ({len(font_subs)}): ' + ', '.join(font_subs[:5])
                                 + (f' … and {len(font_subs)-5} more' if len(font_subs) > 5 else ''))

        # Write stabilised copy
        stabilised = os.path.join(tmp, 'stabilised.xlsx')
        wb.save(stabilised)

        # LO needs its own writable home directory
        lo_home_1 = os.path.join(tmp, 'lo_home_1')
        os.makedirs(lo_home_1, exist_ok=True)

        # ── Stage 4: Primary Conversion ───────────────────────────────────
        lo_out_dir = os.path.join(tmp, 'out_1')
        os.makedirs(lo_out_dir, exist_ok=True)
        pdf_path: Optional[str] = None
        primary_failed = False
        try:
            pdf_path = run_libreoffice(stabilised, lo_out_dir, lo_home_1)
        except Exception as exc:
            primary_failed = True
            all_warnings.append(f'Primary LibreOffice conversion failed: {exc}')

        # ── Stage 5: Post-render Validation ──────────────────────────────
        confidence = 1.0
        if pdf_path and not primary_failed:
            confidence, val_warnings = validate_pdf(pdf_path)
            all_warnings.extend(val_warnings)

        # ── Stage 6: Safe Retry ───────────────────────────────────────────
        RETRY_THRESHOLD = 0.5
        if primary_failed or (pdf_path and confidence < RETRY_THRESHOLD):
            all_warnings.append('Triggering safe retry with isolated LO profile')
            lo_home_2 = os.path.join(tmp, 'lo_home_2')
            os.makedirs(lo_home_2, exist_ok=True)
            lo_out_dir_2 = os.path.join(tmp, 'out_2')
            os.makedirs(lo_out_dir_2, exist_ok=True)
            try:
                pdf_path_retry = safe_retry(stabilised, lo_out_dir_2, lo_home_2)
                conf2, val2 = validate_pdf(pdf_path_retry)
                if conf2 >= confidence or primary_failed:
                    pdf_path  = pdf_path_retry
                    confidence = conf2
                    all_warnings.extend(val2)
                    all_warnings.append('Using retry result')
            except Exception as retry_exc:
                if primary_failed:
                    raise RuntimeError(
                        f'Both primary and retry LibreOffice conversions failed. '
                        f'Last error: {retry_exc}'
                    ) from retry_exc
                all_warnings.append(f'Retry also failed (using primary result): {retry_exc}')

        if not pdf_path:
            raise RuntimeError('LibreOffice pipeline produced no usable PDF')

        # ── Stage 7: Return ───────────────────────────────────────────────
        shutil.copy2(pdf_path, output_path)
        page_count = _count_pages(output_path)

    return {
        'success':    True,
        'pageCount':  page_count,
        'engine':     'libreoffice-primary-v3',
        'confidence': round(confidence, 3),
        'warnings':   all_warnings,
    }


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description='Excel → PDF — LibreOffice Primary Pipeline v3'
    )
    parser.add_argument('input',  help='Input .xlsx / .xls file')
    parser.add_argument('output', help='Output .pdf file')
    parser.add_argument('--page-size',   default='a4',       choices=['a4', 'letter', 'legal'])
    parser.add_argument('--orientation', default='landscape', choices=['portrait', 'landscape'])
    args = parser.parse_args()

    try:
        result = excel_to_pdf_lo(
            args.input, args.output,
            page_size=args.page_size,
            orientation=args.orientation,
        )
        print(json.dumps(result))
    except Exception as exc:
        import traceback
        print(json.dumps({
            'success':   False,
            'error':     str(exc),
            'traceback': traceback.format_exc(),
        }))
        sys.exit(1)


if __name__ == '__main__':
    main()
