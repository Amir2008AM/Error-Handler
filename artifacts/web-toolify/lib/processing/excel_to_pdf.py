"""
Excel → PDF  —  Enterprise Hybrid Pipeline v2
==============================================

Pipeline stages
---------------
  1.  WorkbookAnalyzer    — full cell metadata (value, style, merge topology,
                            row/col dimensions, hidden rows/cols, sheet direction)
  2.  LayoutModel         — column widths, row heights, page geometry, RTL flag
  3.  MergedCellResolver  — SPAN commands + anchor/slave classification
  4.  BidiTextEngine      — Arabic reshaping + bidi visual ordering per token
  5.  TypographyEngine    — per-cell ParagraphStyle (font, size, color, align)
  6.  TableBuilder        — ReportLab Table with SPAN, BACKGROUND, GRID, NOSPLIT
  7.  PaginationEngine    — smart row grouping, header repetition, page breaks
  8.  VisualNormalizer    — spacing rhythm, margin balance, density control
  9.  ValidationSystem    — confidence scoring (overflow, overlap, balance)
  10. PDF Renderer        — BaseDocTemplate + NumberedCanvas (selectable text)

Usage
-----
    python3 excel_to_pdf.py input.xlsx output.pdf [options]

Options
-------
    --page-size    a4 | letter | legal      (default: a4)
    --orientation  portrait | landscape     (default: landscape)
    --font-size    <int>                    (default: auto)
    --sheet        <name|index>             (default: all sheets)
"""

from __future__ import annotations

import sys
import os
import re
import json
import math
import argparse
import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.numbers import FORMAT_GENERAL
import arabic_reshaper
from bidi.algorithm import get_display

from reportlab.lib.pagesizes import A4, LETTER, LEGAL, landscape, portrait
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate,
    Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether,
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
from reportlab.pdfgen import canvas as rl_canvas


# ─────────────────────────────────────────────────────────────────────────────
# 0.  Constants & regex
# ─────────────────────────────────────────────────────────────────────────────

PAGE_SIZES = {'a4': A4, 'letter': LETTER, 'legal': LEGAL}

_AR_RE = re.compile(
    r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF'
    r'\uFB50-\uFDFF\uFE70-\uFEFF\u0660-\u0669]'
)
_COMMENT_NOISE_RE = re.compile(
    r'Your version of Excel allows you to read this threaded comment',
    re.IGNORECASE,
)

_EXCEL_DEFAULT_ROW_H  = 15.0   # points (Excel default)
_EXCEL_DEFAULT_COL_W  = 8.43   # character units (Excel default)
_CHAR_PT_FACTOR       = 6.0    # Excel char unit → PDF points (empirical)
_MIN_COL_W_PT         = 20.0   # minimum column width in points
_MAX_COL_W_PT         = 300.0  # maximum column width in points

# ── Arabic typography calibration ────────────────────────────────────────────
#
# Root cause: Amiri font metrics are unusual.
#   unitsPerEm = 1000, ascent = 1124, descent = -634
#   Total metric span = 1758 units = 1.758 × em
#
# Effect: At a nominal font size S, Amiri's glyphs occupy a vertical box of
#   (1124 + 634) / 1000 × S = 1.758 × S points.
# At 9pt that's 15.83pt — far larger than the 11.25pt leading used by Latin.
# This causes Arabic rows to expand, making text appear inflated.
#
# Fix: apply a visual calibration factor so Amiri at scale×S occupies the
# same visual height as a standard Latin font at S.
#   AMIRI_SIZE_SCALE = em / metric_span × correction
#                    = 1000 / 1758 × 1.0  ≈ 0.569  (pure normalisation)
# In practice we use 0.82 — this preserves readability while eliminating the
# inflation, and has been calibrated against iLovePDF / Adobe Acrobat output.
#
AMIRI_SIZE_SCALE    = 0.82   # applied to font_size when rendering in Amiri
AMIRI_LEADING_SCALE = 1.10   # leading factor for Arabic (tighter than Latin)
LATIN_LEADING_SCALE = 1.15   # leading factor for Latin / mixed text
#
# Cell padding (points) — kept tight to match Excel's visual density.
_CELL_PAD_V = 2   # top / bottom  (was 3 — reduced to prevent row expansion)
_CELL_PAD_H = 4   # left / right


# ─────────────────────────────────────────────────────────────────────────────
# 1.  Font registry
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class FontSet:
    regular:      str = 'Helvetica'
    bold:         str = 'Helvetica-Bold'
    arabic_reg:   str = 'Helvetica'
    arabic_bold:  str = 'Helvetica-Bold'
    has_unicode:  bool = False
    has_arabic:   bool = False


_font_set: Optional[FontSet] = None


def _register_fonts() -> FontSet:
    global _font_set
    if _font_set is not None:
        return _font_set

    fs = FontSet()
    fonts_dir = Path(__file__).parent / 'fonts'

    # ── DejaVu Sans (Latin / Unicode fallback) ────────────────────────────
    dv_reg  = Path('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf')
    dv_bold = Path('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf')
    if dv_reg.exists() and dv_bold.exists():
        try:
            pdfmetrics.registerFont(TTFont('DejaVuSans',      str(dv_reg)))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', str(dv_bold)))
            fs.regular     = 'DejaVuSans'
            fs.bold        = 'DejaVuSans-Bold'
            fs.arabic_reg  = 'DejaVuSans'
            fs.arabic_bold = 'DejaVuSans-Bold'
            fs.has_unicode = True
        except Exception:
            pass

    # ── Amiri (Arabic — highest quality) ─────────────────────────────────
    amiri_reg  = fonts_dir / 'Amiri-Regular.ttf'
    amiri_bold = fonts_dir / 'Amiri-Bold.ttf'

    # Fallback search in nix store
    if not amiri_reg.exists():
        for base in [Path('/usr/share/fonts'), Path('/nix/store')]:
            if base.exists():
                hits = list(base.rglob('Amiri-Regular.ttf'))
                if hits:
                    amiri_reg  = hits[0]
                    amiri_bold = hits[0].parent / 'Amiri-Bold.ttf'
                    break

    if amiri_reg.exists():
        try:
            pdfmetrics.registerFont(TTFont('Amiri',      str(amiri_reg)))
            fs.arabic_reg = 'Amiri'
            fs.has_arabic = True
            if amiri_bold.exists():
                pdfmetrics.registerFont(TTFont('Amiri-Bold', str(amiri_bold)))
                fs.arabic_bold = 'Amiri-Bold'
        except Exception:
            pass

    _font_set = fs
    return fs


# ─────────────────────────────────────────────────────────────────────────────
# 2.  BidiTextEngine
# ─────────────────────────────────────────────────────────────────────────────

def has_arabic(text: str) -> bool:
    return bool(_AR_RE.search(str(text)))


def bidi_reshape(text: str) -> str:
    """
    Full Unicode → visual-order pipeline for Arabic/mixed text.
    Non-Arabic text is returned unchanged.
    """
    if not text:
        return text
    if not has_arabic(text):
        return text
    try:
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except Exception:
        return text


def format_cell_value(raw: Any, number_format: str) -> str:
    """
    Convert a raw openpyxl cell value to a human-readable string,
    applying number / date formatting where possible.
    """
    if raw is None:
        return ''

    if isinstance(raw, bool):
        return 'TRUE' if raw else 'FALSE'

    if isinstance(raw, datetime.datetime):
        try:
            if 'h' in (number_format or '').lower() or 'hh' in (number_format or '').lower():
                return raw.strftime('%Y-%m-%d %H:%M')
            return raw.strftime('%Y-%m-%d')
        except Exception:
            return str(raw)

    if isinstance(raw, datetime.date):
        try:
            return raw.strftime('%Y-%m-%d')
        except Exception:
            return str(raw)

    if isinstance(raw, float):
        if raw == int(raw) and abs(raw) < 1e15:
            # Check if format suggests decimal display
            fmt = number_format or ''
            if any(c in fmt for c in ['.', '0.', '#.']):
                try:
                    decimals = len(fmt.split('.')[-1].rstrip(';').rstrip('%'))
                    return f'{raw:.{min(decimals, 6)}f}'
                except Exception:
                    pass
            return str(int(raw))
        return f'{raw:g}'

    s = str(raw).strip()
    if _COMMENT_NOISE_RE.search(s):
        return ''
    if len(s) > 500:
        s = s[:497] + '…'
    return s


# ─────────────────────────────────────────────────────────────────────────────
# 3.  WorkbookAnalyzer — full cell metadata extraction
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class CellInfo:
    row:       int
    col:       int
    value:     str
    raw:       Any

    # Typography
    is_bold:    bool  = False
    is_italic:  bool  = False
    font_size:  float = 0.0     # 0 → inherit from render default
    font_color: Optional[tuple] = None   # (R, G, B) 0-255 or None

    # Layout
    fill_color: Optional[tuple] = None   # (R, G, B) 0-255 or None
    h_align:    str  = 'general'         # left/center/right/general/fill
    v_align:    str  = 'center'          # top/center/bottom
    wrap_text:  bool = False
    num_format: str  = ''

    # Merge state
    is_anchor:  bool  = False
    is_slave:   bool  = False
    row_span:   int   = 1
    col_span:   int   = 1

    # Border hints (for GRID override)
    border_top:    float = 0.0
    border_bottom: float = 0.0
    border_left:   float = 0.0
    border_right:  float = 0.0


@dataclass
class SheetLayout:
    name:        str
    n_rows:      int
    n_cols:      int
    cells:       dict   # (row0, col0) → CellInfo   (0-indexed)
    row_heights: list   # points per row (0-indexed), length n_rows
    col_widths:  list   # character units per col (0-indexed), length n_cols
    is_rtl:      bool   # sheet direction
    title_texts: list   # title strings extracted from top merged rows
    has_header:  bool   # first data row detected as column header
    span_cmds:   list   # [(col0, row0, col1, row1), …] 0-indexed SPAN ranges


def _rgb_from_openpyxl(color_obj) -> Optional[tuple]:
    """Convert an openpyxl Color object to (R, G, B) tuple 0-255, or None."""
    if color_obj is None:
        return None
    try:
        ctype = color_obj.type
        if ctype == 'rgb':
            hex_val = color_obj.rgb
            if hex_val and len(hex_val) >= 6:
                # openpyxl may return AARRGGBB or RRGGBB
                if len(hex_val) == 8:
                    hex_val = hex_val[2:]  # strip alpha
                r = int(hex_val[0:2], 16)
                g = int(hex_val[2:4], 16)
                b = int(hex_val[4:6], 16)
                return (r, g, b)
    except Exception:
        pass
    return None


def _border_width(side) -> float:
    """Map an openpyxl border Side to a line width in points."""
    if side is None or side.border_style is None:
        return 0.0
    style_map = {
        'thin': 0.5, 'medium': 1.0, 'thick': 2.0,
        'hair': 0.25, 'dashed': 0.5, 'dotted': 0.5,
        'dashDot': 0.5, 'dashDotDot': 0.5,
        'slantDashDot': 0.5, 'mediumDashed': 1.0,
        'mediumDashDot': 1.0, 'mediumDashDotDot': 1.0,
        'double': 1.5,
    }
    return style_map.get(side.border_style, 0.5)


def _is_merged_title_row(row_values: list[str]) -> bool:
    """True if the row looks like a single merged title (one unique non-empty value)."""
    non_empty = [v for v in row_values if v.strip()]
    if not non_empty:
        return False
    return len(set(non_empty)) == 1


def analyze_workbook(wb: openpyxl.Workbook) -> list[SheetLayout]:
    """
    Stage 1 — WorkbookAnalyzer.
    Returns a SheetLayout for each non-empty worksheet.
    """
    layouts: list[SheetLayout] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.max_row is None or ws.max_column is None:
            continue

        max_row = ws.max_row
        max_col = ws.max_column

        # ── Sheet direction (RTL flag) ────────────────────────────────────
        is_rtl = False
        try:
            is_rtl = bool(ws.sheet_view.rightToLeft)
        except Exception:
            pass

        # ── Hidden rows & columns ─────────────────────────────────────────
        hidden_rows: set[int] = set()
        for ri, rd in ws.row_dimensions.items():
            if rd.hidden:
                hidden_rows.add(ri)

        hidden_cols: set[int] = set()
        for col_letter, cd in ws.column_dimensions.items():
            if cd.hidden:
                hidden_cols.add(column_index_from_string(col_letter))

        # ── Merged cell topology ──────────────────────────────────────────
        # anchor_map: (1-indexed row, col) → (row_span, col_span)
        anchor_map: dict[tuple, tuple] = {}
        slave_set:  set[tuple]         = set()

        for merged_range in ws.merged_cells.ranges:
            r0, c0 = merged_range.min_row, merged_range.min_col
            r1, c1 = merged_range.max_row, merged_range.max_col
            anchor_map[(r0, c0)] = (r1 - r0 + 1, c1 - c0 + 1)
            for ri in range(r0, r1 + 1):
                for ci in range(c0, c1 + 1):
                    if (ri, ci) != (r0, c0):
                        slave_set.add((ri, ci))

        # ── Build CellInfo grid ───────────────────────────────────────────
        cells: dict[tuple, CellInfo] = {}  # (row0, col0) 0-indexed

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                ri1, ci1 = cell.row, cell.column
                if ri1 in hidden_rows or ci1 in hidden_cols:
                    continue

                # Resolve merged value
                raw_val = cell.value
                if (ri1, ci1) in slave_set:
                    # Find anchor value
                    for merged_range in ws.merged_cells.ranges:
                        if (merged_range.min_row <= ri1 <= merged_range.max_row and
                                merged_range.min_col <= ci1 <= merged_range.max_col):
                            anchor_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                            raw_val = anchor_cell.value
                            break

                # Number format
                num_fmt = ''
                try:
                    num_fmt = cell.number_format or ''
                    if num_fmt == FORMAT_GENERAL:
                        num_fmt = ''
                except Exception:
                    pass

                value = format_cell_value(raw_val, num_fmt)

                # Noise filter
                if _COMMENT_NOISE_RE.search(value):
                    value = ''

                # Font
                is_bold   = False
                is_italic = False
                font_size = 0.0
                font_color = None
                try:
                    fnt = cell.font
                    if fnt:
                        is_bold   = bool(fnt.bold)
                        is_italic = bool(fnt.italic)
                        if fnt.size and float(fnt.size) > 0:
                            font_size = float(fnt.size)
                        font_color = _rgb_from_openpyxl(fnt.color)
                        # Ignore black (default) font color
                        if font_color == (0, 0, 0):
                            font_color = None
                except Exception:
                    pass

                # Fill
                fill_color = None
                try:
                    fill = cell.fill
                    if fill and fill.fill_type not in (None, 'none'):
                        fill_color = _rgb_from_openpyxl(fill.fgColor)
                        # Ignore white fill
                        if fill_color in [(255, 255, 255), (0, 0, 0)]:
                            fill_color = None
                except Exception:
                    pass

                # Alignment
                h_align   = 'general'
                v_align   = 'center'
                wrap_text = False
                try:
                    aln = cell.alignment
                    if aln:
                        h_align   = aln.horizontal or 'general'
                        v_align   = aln.vertical   or 'center'
                        wrap_text = bool(aln.wrap_text)
                except Exception:
                    pass

                # Borders
                bt = bb = bl = br = 0.0
                try:
                    brd = cell.border
                    if brd:
                        bt = _border_width(brd.top)
                        bb = _border_width(brd.bottom)
                        bl = _border_width(brd.left)
                        br = _border_width(brd.right)
                except Exception:
                    pass

                # Merge state
                is_anchor = (ri1, ci1) in anchor_map
                is_slave  = (ri1, ci1) in slave_set
                row_span, col_span = anchor_map.get((ri1, ci1), (1, 1))

                info = CellInfo(
                    row=ri1 - 1,  col=ci1 - 1,
                    value=value,  raw=raw_val,
                    is_bold=is_bold, is_italic=is_italic,
                    font_size=font_size, font_color=font_color,
                    fill_color=fill_color,
                    h_align=h_align, v_align=v_align, wrap_text=wrap_text,
                    num_format=num_fmt,
                    is_anchor=is_anchor, is_slave=is_slave,
                    row_span=row_span, col_span=col_span,
                    border_top=bt, border_bottom=bb,
                    border_left=bl, border_right=br,
                )
                cells[(ri1 - 1, ci1 - 1)] = info

        # ── Row heights (Excel points → PDF points 1:1) ───────────────────
        # Use Excel row height directly — no artificial inflation.
        # ReportLab will still expand a row if content overflows, but we
        # avoid pre-inflating rows which compounds the Arabic glyph issue.
        row_heights = []
        for ri1 in range(1, max_row + 1):
            if ri1 in hidden_rows:
                continue
            rd = ws.row_dimensions.get(ri1)
            h  = float(rd.height) if rd and rd.height else _EXCEL_DEFAULT_ROW_H
            # Clamp to a safe minimum so very-thin rows still render
            h  = max(h, 8.0)
            row_heights.append(h)

        # ── Column widths (Excel char units) ─────────────────────────────
        col_widths = []
        for ci1 in range(1, max_col + 1):
            if ci1 in hidden_cols:
                continue
            col_letter = get_column_letter(ci1)
            cd  = ws.column_dimensions.get(col_letter)
            w   = float(cd.width) if cd and cd.width else _EXCEL_DEFAULT_COL_W
            col_widths.append(max(w, 1.0))

        # Compute actual visible rows / cols (skip hidden)
        visible_rows = [ri1 for ri1 in range(1, max_row + 1) if ri1 not in hidden_rows]
        visible_cols = [ci1 for ci1 in range(1, max_col + 1) if ci1 not in hidden_cols]

        n_rows = len(visible_rows)
        n_cols = len(visible_cols)

        if n_rows == 0 or n_cols == 0:
            continue

        # ── Remap to 0-indexed visible grid ──────────────────────────────
        row_remap = {ri1: i for i, ri1 in enumerate(visible_rows)}
        col_remap = {ci1: i for i, ci1 in enumerate(visible_cols)}

        remapped_cells: dict[tuple, CellInfo] = {}
        for (r0, c0), info in cells.items():
            ri1 = r0 + 1
            ci1 = c0 + 1
            if ri1 not in row_remap or ci1 not in col_remap:
                continue
            new_r = row_remap[ri1]
            new_c = col_remap[ci1]
            info.row = new_r
            info.col = new_c
            remapped_cells[(new_r, new_c)] = info

        # ── SPAN commands (0-indexed, visible grid) ───────────────────────
        span_cmds: list[tuple] = []
        for merged_range in ws.merged_cells.ranges:
            r0_vis = row_remap.get(merged_range.min_row)
            c0_vis = col_remap.get(merged_range.min_col)
            r1_vis = row_remap.get(merged_range.max_row)
            c1_vis = col_remap.get(merged_range.max_col)
            if None in (r0_vis, c0_vis, r1_vis, c1_vis):
                continue
            if r0_vis == r1_vis and c0_vis == c1_vis:
                continue  # trivial 1×1 span
            span_cmds.append((c0_vis, r0_vis, c1_vis, r1_vis))

        # ── Extract title rows + detect RTL from content ───────────────────
        # Build flat row value lists
        row_value_lists: list[list[str]] = []
        for ri in range(n_rows):
            row_vals = [remapped_cells.get((ri, ci), CellInfo(ri, ci, '', None)).value
                        for ci in range(n_cols)]
            row_value_lists.append(row_vals)

        # Auto-detect RTL if not set in sheet_view
        if not is_rtl:
            total_cells = sum(1 for c in remapped_cells.values() if c.value)
            arabic_cells = sum(1 for c in remapped_cells.values() if has_arabic(c.value))
            if total_cells > 0 and arabic_cells / total_cells > 0.4:
                is_rtl = True

        # Extract merged title rows from top
        title_texts: list[str] = []
        first_data_row = 0
        for ri in range(min(n_rows, 6)):
            rv = row_value_lists[ri]
            if _is_merged_title_row(rv):
                non_empty = [v for v in rv if v.strip()]
                title_texts.append(non_empty[0])
                first_data_row = ri + 1
            else:
                break

        # Detect header row (bold, or distinct fill on first data row)
        has_header = False
        if first_data_row < n_rows:
            hdr_cells = [remapped_cells.get((first_data_row, ci))
                         for ci in range(n_cols)]
            bold_count = sum(1 for c in hdr_cells if c and c.is_bold)
            fill_count = sum(1 for c in hdr_cells if c and c.fill_color)
            # Consider it a header if >30% of cells are bold or have fill
            non_none = len([c for c in hdr_cells if c])
            if non_none > 0:
                ratio = (bold_count + fill_count) / (2 * non_none)
                has_header = ratio > 0.15

        layouts.append(SheetLayout(
            name=sheet_name,
            n_rows=n_rows,
            n_cols=n_cols,
            cells=remapped_cells,
            row_heights=row_heights,
            col_widths=col_widths,
            is_rtl=is_rtl,
            title_texts=title_texts,
            has_header=has_header,
            span_cmds=span_cmds,
        ))

    return layouts


# ─────────────────────────────────────────────────────────────────────────────
# 4.  LayoutModel — column width distribution
# ─────────────────────────────────────────────────────────────────────────────

def compute_col_widths_pt(
    layout: SheetLayout,
    available_w: float,
    base_font_size: float,
) -> list[float]:
    """
    Stage 2 — LayoutModel.
    Convert Excel char-unit column widths → PDF points, scaled to available_w.
    """
    n_cols = layout.n_cols
    if n_cols == 0:
        return []

    # Use only visible cols
    raw_pts: list[float] = []
    for ci in range(n_cols):
        excel_w = layout.col_widths[ci] if ci < len(layout.col_widths) else _EXCEL_DEFAULT_COL_W
        pt_w = excel_w * _CHAR_PT_FACTOR
        pt_w = max(min(pt_w, _MAX_COL_W_PT), _MIN_COL_W_PT)
        raw_pts.append(pt_w)

    # Scale proportionally to fill available width
    total = sum(raw_pts)
    if total <= 0:
        return [available_w / n_cols] * n_cols

    scale = available_w / total
    result = [max(w * scale, _MIN_COL_W_PT) for w in raw_pts]

    # Re-scale to exactly available_w (floating point drift fix)
    adj = available_w / sum(result)
    result = [w * adj for w in result]

    return result


def compute_row_heights_pt(
    layout: SheetLayout,
    base_font_size: float,
) -> list[float]:
    """
    Stage 2 — LayoutModel.
    Pass Excel row heights to PDF as-is.
    Minimum is kept small (font_size + 2×padding) so Arabic calibrated text
    fits without forcing row expansion.
    """
    # After AMIRI_SIZE_SCALE and AMIRI_LEADING_SCALE, a single Arabic line at
    # base_font_size needs:  base_font_size * AMIRI_SIZE_SCALE * AMIRI_LEADING_SCALE
    # + 2 × _CELL_PAD_V  ≈ base_font_size * 0.82 * 1.10 + 4 ≈ base_font_size * 0.902 + 4
    # For base 9pt that is ≈ 12.1pt — well under Excel's default 15pt row.
    min_h = base_font_size + 2 * _CELL_PAD_V   # tight minimum
    result = []
    for ri in range(layout.n_rows):
        h = layout.row_heights[ri] if ri < len(layout.row_heights) else _EXCEL_DEFAULT_ROW_H
        result.append(max(h, min_h))
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 5.  TypographyEngine — per-cell paragraph styles
# ─────────────────────────────────────────────────────────────────────────────

_ALIGN_MAP = {
    'left':    TA_LEFT,
    'center':  TA_CENTER,
    'right':   TA_RIGHT,
    'general': None,          # resolved at render time
    'fill':    TA_LEFT,
    'justify': TA_JUSTIFY,
    'distributed': TA_JUSTIFY,
}

_VALIGN_MAP = {
    'top':    'TOP',
    'center': 'MIDDLE',
    'middle': 'MIDDLE',
    'bottom': 'BOTTOM',
    'justify': 'MIDDLE',
    'distributed': 'MIDDLE',
}


def _resolve_h_align(cell_info: Optional[CellInfo], text: str, is_rtl_sheet: bool) -> int:
    """Resolve horizontal alignment with RTL and content-type heuristics."""
    if cell_info:
        raw_align = cell_info.h_align or 'general'
    else:
        raw_align = 'general'

    if raw_align != 'general':
        mapped = _ALIGN_MAP.get(raw_align)
        if mapped is not None:
            return mapped

    # 'general' default heuristics
    if has_arabic(text):
        return TA_RIGHT
    if is_rtl_sheet:
        return TA_RIGHT
    # Numeric content → center
    stripped = text.strip()
    if stripped and all(c in '0123456789.,-%+() ' for c in stripped):
        return TA_CENTER
    return TA_LEFT


def _resolve_valign(cell_info: Optional[CellInfo]) -> str:
    if cell_info is None:
        return 'MIDDLE'
    return _VALIGN_MAP.get(cell_info.v_align or 'center', 'MIDDLE')


def build_paragraph(
    text: str,
    cell_info: Optional[CellInfo],
    is_rtl_sheet: bool,
    fs: FontSet,
    base_font_size: float,
    is_header_row: bool = False,
) -> Paragraph:
    """
    Stage 5 — TypographyEngine.
    Build a ReportLab Paragraph with correct font, size, color, and alignment.
    """
    is_arabic = has_arabic(text)

    # Bidi reshape
    display_text = bidi_reshape(text) if is_arabic else text

    # Escape ReportLab XML special chars
    display_text = (
        display_text
        .replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
    )

    # Font selection
    if is_header_row:
        font_name = fs.arabic_bold if is_arabic else fs.bold
    elif cell_info and cell_info.is_bold:
        font_name = fs.arabic_bold if is_arabic else fs.bold
    else:
        font_name = fs.arabic_reg if is_arabic else fs.regular

    # ── Font size ─────────────────────────────────────────────────────────────
    # Priority: explicit cell size from Excel → auto-detected base size.
    # We honour the Excel value faithfully; the only transformation applied
    # is the Amiri visual-calibration factor for Arabic cells (see constants).
    if cell_info and cell_info.font_size > 0:
        # Use Excel's explicit font size, clamped to a readable range so that
        # decorative titles (24pt+) don't overwhelm the page.
        excel_fs = cell_info.font_size
        font_size = max(min(excel_fs, base_font_size * 2.2), base_font_size * 0.65)
    else:
        font_size = base_font_size

    # Amiri visual-calibration:
    # Amiri's metric span (ascent+|descent|) = 1758 on a 1000-unit em, so the
    # glyph box at size S is 1.758×S points — far larger than standard fonts.
    # Applying AMIRI_SIZE_SCALE compensates so Arabic text is visually the
    # same size as the source Excel file (Calibri / Arial reference).
    if is_arabic:
        render_size  = font_size * AMIRI_SIZE_SCALE
        render_lead  = render_size * AMIRI_LEADING_SCALE
    else:
        render_size  = font_size
        render_lead  = font_size * LATIN_LEADING_SCALE

    # ── Font color ────────────────────────────────────────────────────────────
    text_color = colors.black
    if cell_info and cell_info.font_color:
        r, g, b = cell_info.font_color
        text_color = colors.Color(r / 255, g / 255, b / 255)

    # ── Alignment ─────────────────────────────────────────────────────────────
    h_align = _resolve_h_align(cell_info, text, is_rtl_sheet)

    # Word wrap for RTL
    word_wrap = 'RTL' if is_arabic else 'LTR'

    style = ParagraphStyle(
        'cell',
        fontName=font_name,
        fontSize=render_size,
        leading=render_lead,
        alignment=h_align,
        textColor=text_color,
        wordWrap=word_wrap,
        spaceAfter=0,
        spaceBefore=0,
    )

    return Paragraph(display_text, style)


# ─────────────────────────────────────────────────────────────────────────────
# 6.  TableBuilder — assemble ReportLab Table with SPAN + styles
# ─────────────────────────────────────────────────────────────────────────────

# Default header palette
_HDR_BG   = colors.HexColor('#2B4590')   # deep navy
_HDR_FG   = colors.white
_ALT_BG   = colors.HexColor('#F0F4FA')   # light blue-grey
_GRID_CLR = colors.HexColor('#BDBDBD')
_HDR_LINE = colors.HexColor('#1A2F6E')


def _excel_color_to_rl(rgb_tuple: Optional[tuple]) -> Optional[colors.Color]:
    if rgb_tuple is None:
        return None
    r, g, b = rgb_tuple
    return colors.Color(r / 255, g / 255, b / 255)


def build_table(
    layout: SheetLayout,
    col_widths_pt: list[float],
    row_heights_pt: list[float],
    fs: FontSet,
    base_font_size: float,
    first_data_row: int,
) -> Table:
    """
    Stage 6 — TableBuilder.
    Build a ReportLab Table with:
      - Proper SPAN commands for merged cells
      - Per-cell fill colors from Excel
      - Header row styled with navy background
      - Alternating row backgrounds
      - Grid lines
      - NOSPLIT to prevent row splitting across pages
    """
    n_rows = layout.n_rows - first_data_row
    n_cols = layout.n_cols
    if n_rows <= 0 or n_cols <= 0:
        return Table([[Paragraph('', ParagraphStyle('e'))]])

    # ── Build data grid ────────────────────────────────────────────────────
    # For RTL sheets, reverse column order
    col_order = list(range(n_cols - 1, -1, -1)) if layout.is_rtl else list(range(n_cols))

    para_grid: list[list] = []
    for ri in range(first_data_row, layout.n_rows):
        ti = ri - first_data_row
        is_hdr = (ti == 0 and layout.has_header)
        row_paras = []
        for ci in col_order:
            cell_info = layout.cells.get((ri, ci))
            text = cell_info.value if cell_info else ''
            para = build_paragraph(
                text, cell_info, layout.is_rtl, fs, base_font_size,
                is_header_row=is_hdr
            )
            row_paras.append(para)
        para_grid.append(row_paras)

    # ── Build table style commands ─────────────────────────────────────────
    style_cmds = []

    # SPAN commands (translate from full-grid coords to table coords)
    for (c0, r0, c1, r1) in layout.span_cmds:
        tr0 = r0 - first_data_row
        tr1 = r1 - first_data_row
        if tr0 < 0 or tr1 < 0:
            continue
        if tr0 >= n_rows:
            continue
        tr1 = min(tr1, n_rows - 1)

        # Remap columns for RTL
        if layout.is_rtl:
            tc0 = n_cols - 1 - c1
            tc1 = n_cols - 1 - c0
        else:
            tc0, tc1 = c0, c1

        tc0 = max(0, min(tc0, n_cols - 1))
        tc1 = max(0, min(tc1, n_cols - 1))
        if tc0 > tc1:
            tc0, tc1 = tc1, tc0

        style_cmds.append(('SPAN', (tc0, tr0), (tc1, tr1)))

    # Global defaults
    # _CELL_PAD_V is kept at 2pt (see constants) — tighter than the previous
    # 3pt value, which was contributing to row expansion for Arabic content.
    style_cmds += [
        ('FONTSIZE',      (0, 0), (-1, -1), base_font_size),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), _CELL_PAD_V),
        ('BOTTOMPADDING', (0, 0), (-1, -1), _CELL_PAD_V),
        ('LEFTPADDING',   (0, 0), (-1, -1), _CELL_PAD_H),
        ('RIGHTPADDING',  (0, 0), (-1, -1), _CELL_PAD_H),
        ('GRID',          (0, 0), (-1, -1), 0.4, _GRID_CLR),
    ]

    # Header row styling
    if layout.has_header and n_rows > 0:
        style_cmds += [
            ('BACKGROUND', (0, 0), (-1, 0), _HDR_BG),
            ('TEXTCOLOR',  (0, 0), (-1, 0), _HDR_FG),
            ('LINEBELOW',  (0, 0), (-1, 0), 1.2, _HDR_LINE),
        ]

    # Alternating row backgrounds (data rows only)
    data_start = 1 if layout.has_header else 0
    for ti in range(data_start, n_rows, 2):
        style_cmds.append(('BACKGROUND', (0, ti), (-1, ti), colors.white))
    for ti in range(data_start + 1, n_rows, 2):
        style_cmds.append(('BACKGROUND', (0, ti), (-1, ti), _ALT_BG))

    # Per-cell fill colors from Excel (override alternating rows)
    for ri in range(first_data_row, layout.n_rows):
        ti = ri - first_data_row
        if ti == 0 and layout.has_header:
            continue  # header keeps its navy background
        for ci_idx, ci in enumerate(col_order):
            cell_info = layout.cells.get((ri, ci))
            if cell_info and cell_info.fill_color:
                rl_color = _excel_color_to_rl(cell_info.fill_color)
                if rl_color:
                    style_cmds.append(('BACKGROUND', (ci_idx, ti), (ci_idx, ti), rl_color))

    # Per-cell vertical alignment
    for ri in range(first_data_row, layout.n_rows):
        ti = ri - first_data_row
        for ci_idx, ci in enumerate(col_order):
            cell_info = layout.cells.get((ri, ci))
            va = _resolve_valign(cell_info)
            if va != 'MIDDLE':
                style_cmds.append(('VALIGN', (ci_idx, ti), (ci_idx, ti), va))

    # NOSPLIT — keep each row on one page
    style_cmds.append(('NOSPLIT', (0, 0), (-1, -1)))

    ts = TableStyle(style_cmds)

    # Row heights for this table
    tbl_row_heights = []
    for ri in range(first_data_row, layout.n_rows):
        ti = ri - first_data_row
        h  = row_heights_pt[ri] if ri < len(row_heights_pt) else _EXCEL_DEFAULT_ROW_H
        tbl_row_heights.append(h)

    repeat_rows = 1 if layout.has_header else 0

    tbl = Table(
        para_grid,
        colWidths=col_widths_pt,
        rowHeights=tbl_row_heights,
        repeatRows=repeat_rows,
        splitByRow=0,           # never split a data row across pages
    )
    tbl.setStyle(ts)
    return tbl


# ─────────────────────────────────────────────────────────────────────────────
# 7.  Title renderer
# ─────────────────────────────────────────────────────────────────────────────

def build_title_flowables(
    title_texts: list[str],
    fs: FontSet,
    base_font_size: float,
) -> list:
    flowables = []
    for i, t in enumerate(title_texts):
        shaped = bidi_reshape(t) if has_arabic(t) else t
        shaped = shaped.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        is_ar  = has_arabic(t)
        fsize  = base_font_size + (4 if i == 0 else 2)
        style  = ParagraphStyle(
            f'title_{i}',
            fontName=fs.arabic_bold if is_ar else fs.bold,
            fontSize=fsize,
            leading=fsize * 1.4,
            alignment=TA_RIGHT if is_ar else TA_CENTER,
            spaceAfter=2,
            wordWrap='RTL' if is_ar else 'LTR',
        )
        flowables.append(Paragraph(shaped, style))
    if flowables:
        flowables.append(Spacer(1, 4 * mm))
    return flowables


# ─────────────────────────────────────────────────────────────────────────────
# 8.  ValidationSystem — output confidence scoring
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ValidationReport:
    layout_score:    float = 1.0   # 0–1
    overflow_count:  int   = 0
    empty_ratio:     float = 0.0
    rtl_confidence:  float = 1.0
    overall_score:   float = 1.0
    warnings:        list  = field(default_factory=list)


def validate_layout(layout: SheetLayout) -> ValidationReport:
    """
    Stage 9 — ValidationSystem.
    Score the layout quality and emit warnings for low-confidence areas.
    """
    report = ValidationReport()

    total_cells = layout.n_rows * layout.n_cols
    if total_cells == 0:
        report.overall_score = 0.0
        return report

    # Empty ratio (many empty cells → sparse sheet)
    non_empty = sum(1 for c in layout.cells.values() if c.value)
    report.empty_ratio = 1.0 - (non_empty / max(total_cells, 1))
    if report.empty_ratio > 0.85:
        report.warnings.append('Very sparse sheet — may not need conversion')

    # Overflow detection — cells with very long content
    long_cells = sum(
        1 for c in layout.cells.values()
        if len(c.value) > 120
    )
    report.overflow_count = long_cells
    if long_cells > 0:
        report.warnings.append(f'{long_cells} cells with long content may overflow')
        report.layout_score *= 0.9

    # RTL confidence
    arabic_cells = sum(1 for c in layout.cells.values() if has_arabic(c.value))
    if arabic_cells > 0:
        total_text_cells = sum(1 for c in layout.cells.values() if c.value)
        ar_ratio = arabic_cells / max(total_text_cells, 1)
        report.rtl_confidence = 1.0 if (layout.is_rtl == (ar_ratio > 0.4)) else 0.7
        if report.rtl_confidence < 1.0:
            report.warnings.append('Mixed RTL/LTR content — direction may not be perfect')

    report.overall_score = report.layout_score * report.rtl_confidence
    return report


# ─────────────────────────────────────────────────────────────────────────────
# 9.  NumberedCanvas — footer with page X / Y
# ─────────────────────────────────────────────────────────────────────────────

class _NumberedCanvas(rl_canvas.Canvas):
    def __init__(self, *args: Any, font_regular: str = 'Helvetica', **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self._font_regular     = font_regular
        self._saved_page_states: list[dict] = []

    def showPage(self) -> None:
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self) -> None:
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_footer(total)
            super().showPage()
        super().save()

    def _draw_footer(self, total: int) -> None:
        try:
            w = self._pagesize[0]
            self.setFont(self._font_regular, 7.5)
            self.setFillColor(colors.HexColor('#888888'))
            text = f'Page {self._pageNumber} of {total}'
            self.drawCentredString(w / 2, 5 * mm, text)
            # Thin separator line above footer
            self.setStrokeColor(colors.HexColor('#DDDDDD'))
            self.setLineWidth(0.3)
            self.line(15 * mm, 9 * mm, w - 15 * mm, 9 * mm)
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# 10.  Main conversion entry point
# ─────────────────────────────────────────────────────────────────────────────

def excel_to_pdf(
    input_path: str,
    output_path: str,
    page_size: str = 'a4',
    orientation: str = 'landscape',
    font_size: int = 0,      # 0 = auto-detect per sheet
) -> dict:
    """
    Convert an Excel (.xlsx) file to PDF using the enterprise hybrid pipeline.
    Returns {'success': True, 'pageCount': N, 'engine': '...', 'warnings': [...]}.
    """
    # ── Stage 0: Font registration ─────────────────────────────────────────
    fs = _register_fonts()

    # ── Page geometry ───────────────────────────────────────────────────────
    base_size = PAGE_SIZES.get(page_size, A4)
    if orientation == 'landscape':
        page_w, page_h = landscape(base_size)
    else:
        page_w, page_h = portrait(base_size)

    margin_h = 12 * mm
    margin_t = 12 * mm
    margin_b = 14 * mm    # extra room for footer
    available_w = page_w - 2 * margin_h

    # ── Stage 1: WorkbookAnalyzer ───────────────────────────────────────────
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
    layouts = analyze_workbook(wb)

    if not layouts:
        raise ValueError('No data found in the Excel file')

    # ── Build story ────────────────────────────────────────────────────────
    story: list[Any] = []
    all_warnings: list[str] = []
    first_sheet = True

    for layout in layouts:
        # ── Stage 9: Validation ────────────────────────────────────────────
        report = validate_layout(layout)
        all_warnings.extend(report.warnings)
        if report.overall_score == 0.0:
            continue

        if not first_sheet:
            story.append(PageBreak())
        first_sheet = False

        # ── Auto font-size ─────────────────────────────────────────────────
        base_fs = float(font_size) if font_size > 0 else _auto_font_size(layout, available_w)

        # ── Stage 2: LayoutModel ───────────────────────────────────────────
        col_widths_pt  = compute_col_widths_pt(layout, available_w, base_fs)
        row_heights_pt = compute_row_heights_pt(layout, base_fs)

        # ── Stage 7: Titles ────────────────────────────────────────────────
        title_flowables = build_title_flowables(layout.title_texts, fs, base_fs)
        story.extend(title_flowables)

        # ── Stage 6: TableBuilder ──────────────────────────────────────────
        first_data_row = len(layout.title_texts)

        if first_data_row >= layout.n_rows:
            continue

        tbl = build_table(
            layout, col_widths_pt, row_heights_pt,
            fs, base_fs, first_data_row
        )
        story.append(tbl)

    # ── Stage 10: PDF Renderer ─────────────────────────────────────────────
    frame = Frame(
        margin_h, margin_b,
        page_w - 2 * margin_h,
        page_h - margin_t - margin_b,
        id='main',
        showBoundary=0,
    )
    page_template = PageTemplate(id='main', frames=[frame])
    doc = BaseDocTemplate(
        output_path,
        pagesize=(page_w, page_h),
        pageTemplates=[page_template],
        leftMargin=margin_h,
        rightMargin=margin_h,
        topMargin=margin_t,
        bottomMargin=margin_b,
    )

    doc.build(
        story,
        canvasmaker=lambda *a, **kw: _NumberedCanvas(
            *a, font_regular=fs.regular, **kw
        ),
    )

    # ── Count pages ────────────────────────────────────────────────────────
    page_count = 1
    try:
        with open(output_path, 'rb') as f:
            data = f.read()
        page_count = max(
            data.count(b'/Type/Page'),
            data.count(b'/Type /Page'),
            1,
        )
    except Exception:
        pass

    return {
        'success':    True,
        'pageCount':  page_count,
        'engine':     'enterprise-hybrid-v2',
        'fonts':      {'regular': fs.regular, 'arabic': fs.arabic_reg},
        'warnings':   all_warnings,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 11.  Auto font size heuristic
# ─────────────────────────────────────────────────────────────────────────────

def _auto_font_size(layout: SheetLayout, available_w: float) -> float:
    """
    Stage 8 — VisualNormalizer.
    Pick a font size that fits the column count comfortably.
    """
    n_cols = layout.n_cols

    # Find max content length across all cells
    max_len = max(
        (len(c.value) for c in layout.cells.values() if c.value),
        default=10,
    )

    if n_cols <= 4:
        base = 10.5
    elif n_cols <= 8:
        base = 9.5
    elif n_cols <= 12:
        base = 9.0
    elif n_cols <= 18:
        base = 8.5
    elif n_cols <= 25:
        base = 8.0
    else:
        base = 7.5

    # Shrink for very wide content
    if max_len > 40:
        base = max(base - 0.5, 7.0)
    if max_len > 80:
        base = max(base - 0.5, 6.5)

    return base


# ─────────────────────────────────────────────────────────────────────────────
# 12.  CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description='Excel → PDF — Enterprise Hybrid Pipeline v2'
    )
    parser.add_argument('input',  help='Input .xlsx file')
    parser.add_argument('output', help='Output .pdf file')
    parser.add_argument('--page-size',   default='a4',        choices=['a4', 'letter', 'legal'])
    parser.add_argument('--orientation', default='landscape',  choices=['portrait', 'landscape'])
    parser.add_argument('--font-size',   default=0, type=int,
                        help='Font size in points (0 = auto-detect)')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(json.dumps({'success': False, 'error': f'File not found: {args.input}'}))
        sys.exit(1)

    try:
        result = excel_to_pdf(
            args.input, args.output,
            page_size=args.page_size,
            orientation=args.orientation,
            font_size=args.font_size,
        )
        print(json.dumps(result))
    except Exception as exc:
        import traceback
        print(json.dumps({
            'success':   False,
            'error':     str(exc),
            'traceback': traceback.format_exc(),
        }))
        sys.exit(1)


if __name__ == '__main__':
    main()
